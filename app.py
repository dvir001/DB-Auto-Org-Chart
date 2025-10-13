from flask import Flask, render_template, render_template_string, jsonify, request, send_from_directory, send_file, session, redirect, url_for
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_session import Session
import atexit
import json
import os
from datetime import datetime, timedelta, timezone
import requests
import threading
import time
from io import BytesIO
import schedule
import logging
from dotenv import load_dotenv
from werkzeug.utils import secure_filename

from functools import wraps
import hashlib
import secrets
import re
try:
    from PIL import Image
except ImportError:
    Image = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

_allowed_origins = [origin.strip() for origin in os.environ.get('CORS_ALLOWED_ORIGINS', '').split(',') if origin.strip()]
if _allowed_origins:
    CORS(app, resources={r"/api/*": {"origins": _allowed_origins}})

# Security Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False

# Initialize extensions
Session(app)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Simple authentication settings
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD')
if not ADMIN_PASSWORD:
    raise RuntimeError('ADMIN_PASSWORD environment variable must be set to a strong value')
if ADMIN_PASSWORD in {'admin123', 'your-admin-password-here'}:
    raise RuntimeError('ADMIN_PASSWORD must not use the default placeholder value')

# Security headers
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'; img-src 'self' data:;"
    return response

# Authentication decorators
def require_auth(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('authenticated'):
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('authenticated'):
            desired_path = sanitize_next_path(request.path)
            params = {'next': desired_path} if desired_path else {}
            return redirect(url_for('login', **params))
        return f(*args, **kwargs)
    return decorated_function


_next_path_pattern = re.compile(r"^[A-Za-z0-9_\-/]*$")


def sanitize_next_path(raw_value):
    if not raw_value:
        return ''

    candidate = raw_value.strip()

    if candidate.startswith(('http://', 'https://', '//')):
        return ''

    candidate = candidate.lstrip('/')

    if not _next_path_pattern.fullmatch(candidate):
        return ''

    return candidate


# Utility functions
def validate_image_file(file):
    """Validate that uploaded file is a safe image"""
    if not Image:
        logger.warning("PIL not available, skipping image validation")
        return True

    try:
        # Open and verify image
        img = Image.open(file)
        img.verify()

        # Reset file pointer
        file.seek(0)

        # Check image dimensions (reasonable limits)
        img = Image.open(file)
        if img.width > 2000 or img.height > 2000:
            return False

        # Reset file pointer again
        file.seek(0)
        return True
    except Exception as e:
        logger.error(f"Image validation failed: {e}")
        return False

if not os.path.exists('static'):
    os.makedirs('static')

GRAPH_API_ENDPOINT = 'https://graph.microsoft.com/v1.0'
GRAPH_API_BETA_ENDPOINT = 'https://graph.microsoft.com/beta'
# DATA_FILE will be set after DATA_DIR is created
# Create data directory for persistent storage
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
logger.info(f"DATA_DIR set to: {DATA_DIR}")
if not os.path.exists(DATA_DIR):
    try:
        os.makedirs(DATA_DIR)
        logger.info(f"Created data directory: {DATA_DIR}")
    except Exception as e:
        logger.warning(f"Could not create data directory: {e}")
        DATA_DIR = SCRIPT_DIR

SETTINGS_FILE = os.path.join(DATA_DIR, 'app_settings.json')
DATA_FILE = os.path.join(DATA_DIR, 'employee_data.json')
MISSING_MANAGER_FILE = os.path.join(DATA_DIR, 'missing_manager_records.json')
EMPLOYEE_LIST_FILE = os.path.join(DATA_DIR, 'employee_list.json')
DISABLED_LICENSE_FILE = os.path.join(DATA_DIR, 'disabled_with_license_records.json')
FILTERED_LICENSE_FILE = os.path.join(DATA_DIR, 'filtered_with_license_records.json')
FILTERED_USERS_FILE = os.path.join(DATA_DIR, 'filtered_user_records.json')
DISABLED_USERS_FILE = os.path.join(DATA_DIR, 'disabled_user_records.json')
LAST_LOGIN_FILE = os.path.join(DATA_DIR, 'last_login_records.json')
RECENTLY_DISABLED_FILE = os.path.join(DATA_DIR, 'recently_disabled_employees.json')
RECENTLY_HIRED_FILE = os.path.join(DATA_DIR, 'recently_hired_employees.json')

# Configuration for file uploads (removed SVG for security)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB limit for logo uploads

TENANT_ID = os.environ.get('AZURE_TENANT_ID')
CLIENT_ID = os.environ.get('AZURE_CLIENT_ID')
CLIENT_SECRET = os.environ.get('AZURE_CLIENT_SECRET')

if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
    logger.warning("Missing Azure AD credentials in environment variables!")
    logger.warning("AZURE_TENANT_ID: " + ("Set" if TENANT_ID else "Not set"))
    logger.warning("AZURE_CLIENT_ID: " + ("Set" if CLIENT_ID else "Not set"))
    logger.warning("AZURE_CLIENT_SECRET: " + ("Set" if CLIENT_SECRET else "Not set"))
    logger.warning("Please check your .env file exists and contains the correct values")

TOP_LEVEL_USER_EMAIL = os.environ.get('TOP_LEVEL_USER_EMAIL')
TOP_LEVEL_USER_ID = os.environ.get('TOP_LEVEL_USER_ID')

scheduler_running = False
scheduler_lock = threading.Lock()

# Default settings
DEFAULT_SETTINGS = {
    'chartTitle': 'DB Auto Org Chart',
    'headerColor': '#0078d4',
    'logoPath': '/static/icon.png',
    'faviconPath': '/favicon.ico',
    'nodeColors': {
        'level0': '#90EE90',
        'level1': '#FFFFE0',
        'level2': '#E0F2FF',
        'level3': '#FFE4E1',
        'level4': '#E8DFF5',
        'level5': '#FFEAA7'
    },
    'autoUpdateEnabled': True,
    'updateTime': '20:00',
    'collapseLevel': '2',
    'searchAutoExpand': True,
    'searchHighlight': True,
    'showNames': True,
    'showDepartments': True,
    'showJobTitles': True,
    'showEmployeeCount': True,
    'showProfileImages': True,
    'printOrientation': 'landscape',
    'printSize': 'a4',
    'exportXlsxColumns': {
        'name': 'show',
        'title': 'show',
        'department': 'show',
        'email': 'show',
        'phone': 'show',
        'businessPhone': 'show',
        'hireDate': 'admin',
        'country': 'show',
        'state': 'show',
        'city': 'show',
        'office': 'show',
        'manager': 'show'
    },
    'topUserEmail': TOP_LEVEL_USER_EMAIL or '',
    'highlightNewEmployees': True,
    'newEmployeeMonths': 3,
    # Multi-line children layout for large teams
    'multiLineChildrenEnabled': True,
    'multiLineChildrenThreshold': 20,
    # Use equal sibling spacing to avoid phantom gaps
    'compactSiblingSpacingEnabled': False,
    'hideDisabledUsers': True,
    'hideGuestUsers': True,
    'hideNoTitle': True,
    'hideConsultantGroup': True,
    # Comma-separated list of department names to ignore (case-insensitive)
    'ignoredEmployees': '',
    'ignoredDepartments': 'Consultant Group',
    # Comma-separated list of job titles to ignore (case-insensitive, substring match)
    'ignoredTitles': ''
}


def translate_placeholder(key, default=None, **kwargs):
    """Basic translation helper used by templates until full i18n is wired."""
    if default is not None:
        try:
            return default.format(**kwargs)
        except Exception:
            return default
    return key

def _apply_environment_overrides(settings):
    settings = settings.copy()

    if TOP_LEVEL_USER_EMAIL:
        settings['topUserEmail'] = TOP_LEVEL_USER_EMAIL.strip()
    else:
        settings['topUserEmail'] = settings.get('topUserEmail', '')

    return settings


def load_settings():
    """Load settings from file or return defaults with environment overrides"""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                settings = json.load(f)
                for key in DEFAULT_SETTINGS:
                    if key not in settings:
                        settings[key] = DEFAULT_SETTINGS[key]
                return _apply_environment_overrides(settings)
        except Exception as e:
            logger.error(f"Error loading settings: {e}")
    defaults = DEFAULT_SETTINGS.copy()
    return _apply_environment_overrides(defaults)

# Helpers: ignored titles/departments parsing and matching
_filter_legacy_split_re = re.compile(r"\s*[;,]+\s*")
_trim_edge_punct = re.compile(r"^[\s\-–—|]+|[\s\-–—|]+$")

def normalize_filter_value(value):
    if not value:
        return ''
    cleaned = _trim_edge_punct.sub('', str(value))
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip().lower()

def parse_filter_values(raw_value):
    if raw_value is None:
        return set()

    values = None

    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return set()
        if text.startswith('['):
            try:
                decoded = json.loads(text)
                if isinstance(decoded, (list, tuple, set)):
                    values = list(decoded)
            except json.JSONDecodeError:
                values = None
        if values is None:
            values = _filter_legacy_split_re.split(text)
    elif isinstance(raw_value, (list, tuple, set)):
        values = list(raw_value)
    else:
        return set()

    normalized = set()
    for part in values:
        normalized_value = normalize_filter_value(part)
        if normalized_value:
            normalized.add(normalized_value)
    return normalized

def parse_ignored_departments(settings):
    raw = settings.get('ignoredDepartments', '')
    return parse_filter_values(raw)

def parse_ignored_titles(settings):
    raw = settings.get('ignoredTitles', '')
    return parse_filter_values(raw)

def parse_ignored_employees(settings):
    raw = settings.get('ignoredEmployees', '')
    return parse_filter_values(raw)

def department_is_ignored(department, ignored_set):
    if not ignored_set:
        return False
    normalized = normalize_filter_value(department)
    return normalized in ignored_set


def employee_is_ignored(name, email, user_principal_name, ignored_values):
    if not ignored_values:
        return False

    candidates = set()

    for value in (name, email, user_principal_name):
        normalized = normalize_filter_value(value)
        if normalized:
            candidates.add(normalized)

    contact_values = [value for value in (email, user_principal_name) if value]
    for contact in contact_values:
        if name:
            for combo in (
                f"{name} <{contact}>",
                f"{name} ({contact})",
                f"{name} - {contact}",
                f"{contact} ({name})",
                f"{contact} - {name}"
            ):
                normalized = normalize_filter_value(combo)
                if normalized:
                    candidates.add(normalized)

    return any(candidate in ignored_values for candidate in candidates)


def parse_graph_datetime(value):
    if not value:
        return None

    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, (int, float)):
        try:
            dt = datetime.fromtimestamp(value, tz=timezone.utc)
        except Exception:
            return None
    elif isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            if text.endswith('Z'):
                text = text[:-1] + '+00:00'
            dt = datetime.fromisoformat(text)
        except ValueError:
            dt = None
            for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
                try:
                    dt = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    dt = None
            if dt is None:
                return None
    else:
        return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    return dt


def datetime_to_iso(dt):
    if not isinstance(dt, datetime):
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).isoformat()


def calculate_days_since(moment):
    dt = parse_graph_datetime(moment)
    if not dt:
        return None
    now = datetime.now(timezone.utc)
    delta = now - dt.astimezone(timezone.utc)
    return max(delta.days, 0)


def collect_recently_disabled_employees(records, days=365):
    if not records:
        return []

    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    recent = []

    for record in records:
        observed_value = (
            record.get('firstSeenDisabledAt')
            or record.get('disabledDate')
        )
        disabled_at = parse_graph_datetime(observed_value)
        if not disabled_at or disabled_at < cutoff:
            continue

        updated = record.copy()
        updated['disabledDate'] = datetime_to_iso(disabled_at)
        updated['disabledDays'] = calculate_days_since(disabled_at)
        if not updated.get('firstSeenDisabledAt'):
            updated['firstSeenDisabledAt'] = updated['disabledDate']
        recent.append(updated)

    recent.sort(key=lambda item: item.get('disabledDate') or '')
    return recent


def collect_recently_hired_employees(employees, days=365):
    if not employees:
        return []

    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    manager_lookup = {emp.get('id'): emp for emp in employees if emp.get('id')}
    recent = []

    for employee in employees:
        hire_date = parse_graph_datetime(employee.get('hireDate') or employee.get('employeeHireDate'))
        if not hire_date or hire_date < cutoff:
            continue

        record = {
            'id': employee.get('id'),
            'name': employee.get('name'),
            'title': employee.get('title'),
            'department': employee.get('department'),
            'email': employee.get('email'),
            'userPrincipalName': employee.get('userPrincipalName'),
            'phone': employee.get('phone') or '',
            'businessPhone': employee.get('businessPhone') or '',
            'location': employee.get('location') or employee.get('officeLocation') or '',
            'hireDate': datetime_to_iso(hire_date),
            'daysSinceHire': calculate_days_since(hire_date),
            'managerName': '',
        }

        manager_id = employee.get('managerId')
        if manager_id and manager_id in manager_lookup:
            record['managerName'] = manager_lookup[manager_id].get('name') or ''

        recent.append(record)

    recent.sort(key=lambda item: item.get('hireDate') or '')
    return recent


def save_settings(settings):
    """Save settings to file"""
    try:
        # Log current working directory and file path for debugging
        logger.info(f"Attempting to save settings to: {SETTINGS_FILE}")
        logger.info(f"Current working directory: {os.getcwd()}")
        logger.info(f"Settings file exists before save: {os.path.exists(SETTINGS_FILE)}")
        
        # Ensure directory exists
        settings_dir = os.path.dirname(SETTINGS_FILE) or '.'
        if not os.path.exists(settings_dir):
            os.makedirs(settings_dir)
        
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(settings, f, indent=2)
        
        logger.info(f"Settings saved successfully. File exists after save: {os.path.exists(SETTINGS_FILE)}")
        return True
    except Exception as e:
        logger.error(f"Error saving settings to {SETTINGS_FILE}: {e}")
        logger.error(f"Current working directory: {os.getcwd()}")
        logger.error(f"Directory permissions: {oct(os.stat('.').st_mode)[-3:] if os.path.exists('.') else 'N/A'}")
        return False

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_access_token():
    token_url = f'https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token'
    
    token_data = {
        'grant_type': 'client_credentials',
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'scope': 'https://graph.microsoft.com/.default'
    }
    
    try:
        token_r = requests.post(token_url, data=token_data, timeout=10)
        token = token_r.json().get('access_token')
        return token
    except Exception as e:
        logger.error(f"Error getting access token: {e}")
        return None

def fetch_employee_photo(user_id, token):
    """Fetch employee photo from Microsoft Graph API"""
    try:
        photo_url = f'{GRAPH_API_ENDPOINT}/users/{user_id}/photo/$value'
        headers = {
            'Authorization': f'Bearer {token}'
        }
        
        response = requests.get(photo_url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            return response.content
        else:
            logger.debug(f"No photo found for user {user_id}: {response.status_code}")
            return None
            
    except Exception as e:
        logger.debug(f"Error fetching photo for user {user_id}: {e}")
        return None

def fetch_all_employees(token=None):
    token = token or get_access_token()
    if not token:
        logger.error("Failed to get access token")
        logger.warning("Using cached employee data because access token retrieval failed")
        return _load_fetch_all_employees_fallback()

    # Load settings to check filtering preferences
    settings = load_settings()
    hide_disabled_users = settings.get('hideDisabledUsers', True)
    hide_guest_users = settings.get('hideGuestUsers', True)
    hide_no_title = settings.get('hideNoTitle', True)
    ignored_title_values = parse_ignored_titles(settings)
    ignored_employee_values = parse_ignored_employees(settings)
    ignored_department_values = parse_ignored_departments(settings)
    new_employee_months = settings.get('newEmployeeMonths', 3)

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    employees = []
    filtered_with_license = []
    filtered_users = []
    fetch_failed = False

    sku_map = fetch_subscribed_sku_map(token)

    # Build API filter based on settings
    api_filters = []
    if hide_disabled_users:
        api_filters.append("accountEnabled eq true")
    if hide_guest_users:
        api_filters.append("userType eq 'Member'")
    
    # Construct the users URL with conditional filtering
    filter_string = " and ".join(api_filters) if api_filters else ""
    select_fields = (
        'id,displayName,jobTitle,department,mail,userPrincipalName,mobilePhone,'
        'businessPhones,officeLocation,city,state,country,usageLocation,streetAddress,'
        'postalCode,employeeHireDate,accountEnabled,userType,assignedLicenses'
    )

    if filter_string:
        users_url = (
            f'{GRAPH_API_ENDPOINT}/users?$select={select_fields}'
            f'&$expand=manager($select=id,displayName)&$filter={filter_string}'
        )
    else:
        users_url = (
            f'{GRAPH_API_ENDPOINT}/users?$select={select_fields}'
            f'&$expand=manager($select=id,displayName)'
        )
    
    while users_url:
        try:
            response = requests.get(users_url, headers=headers, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if 'value' in data:
                for user in data['value']:
                    display_name = user.get('displayName') or ''
                    primary_email = user.get('mail') or ''
                    user_principal_name = user.get('userPrincipalName') or ''
                    job_title_val = user.get('jobTitle') or ''
                    lowered_title = normalize_filter_value(job_title_val)
                    department_val = user.get('department') or ''
                    business_phones = user.get('businessPhones') or []
                    if isinstance(business_phones, list):
                        business_phone = next((phone for phone in business_phones if phone), '')
                    else:
                        business_phone = business_phones or ''

                    assigned_licenses = user.get('assignedLicenses') or []
                    user_type = (user.get('userType') or '').lower()

                    license_sku_ids = []
                    license_labels = []
                    if assigned_licenses:
                        seen_labels = set()
                        for license_entry in assigned_licenses:
                            sku_id = license_entry.get('skuId')
                            if not sku_id:
                                continue
                            sku_key = str(sku_id).lower()
                            license_sku_ids.append(str(sku_id))
                            friendly_name = (
                                sku_map.get(sku_key)
                                or sku_map.get(sku_key.upper())
                                or str(sku_id)
                            )
                            normalized_label = friendly_name.lower()
                            if normalized_label not in seen_labels:
                                seen_labels.add(normalized_label)
                                license_labels.append(friendly_name)

                        license_labels.sort(key=lambda item: item.lower())

                    filtered_reasons = []
                    if hide_disabled_users and not user.get('accountEnabled', True):
                        filtered_reasons.append('filter_disabled')
                    if hide_guest_users and user_type == 'guest':
                        filtered_reasons.append('filter_guest')
                    if hide_no_title and job_title_val.strip() == '':
                        filtered_reasons.append('filter_no_title')
                    if ignored_title_values and lowered_title in ignored_title_values:
                        filtered_reasons.append('filter_ignored_title')
                    if department_is_ignored(department_val, ignored_department_values):
                        filtered_reasons.append('filter_ignored_department')
                    if employee_is_ignored(display_name, primary_email, user_principal_name, ignored_employee_values):
                        filtered_reasons.append('filter_ignored_employee')

                    if filtered_reasons:
                        base_record = {
                            'id': user.get('id'),
                            'name': display_name or 'Unknown',
                            'title': job_title_val or 'No Title',
                            'department': department_val or 'No Department',
                            'email': primary_email or user_principal_name or '',
                            'userPrincipalName': user_principal_name,
                            'phone': user.get('mobilePhone') or '',
                            'businessPhone': business_phone,
                            'location': user.get('officeLocation') or '',
                            'city': user.get('city') or '',
                            'state': user.get('state') or '',
                            'country': user.get('country') or '',
                            'usageLocation': user.get('usageLocation') or '',
                            'accountEnabled': user.get('accountEnabled', True),
                            'userType': user_type,
                            'filterReasons': filtered_reasons,
                            'licenseCount': len(license_sku_ids),
                            'licenseSkus': license_labels,
                            'licenseSkuIds': license_sku_ids,
                        }
                        filtered_users.append(base_record)

                        if license_sku_ids:
                            filtered_with_license.append(dict(base_record))
                        continue

                    if display_name:
                        hire_date_str = user.get('employeeHireDate')
                        is_new = False
                        hire_date = None

                        if hire_date_str:
                            try:
                                if 'T' in hire_date_str:
                                    hire_date = datetime.fromisoformat(hire_date_str.replace('Z', '+00:00'))
                                else:
                                    hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d')
                                    hire_date = hire_date.replace(tzinfo=None)

                                if hire_date.tzinfo:
                                    cutoff_date = datetime.now(hire_date.tzinfo) - timedelta(days=new_employee_months * 30)
                                else:
                                    cutoff_date = datetime.now() - timedelta(days=new_employee_months * 30)

                                is_new = hire_date > cutoff_date
                            except Exception as e:
                                logger.warning(f"Error parsing hire date for user {user.get('displayName')}: {e}")

                        # Build full address from components
                        address_components = []
                        if user.get('streetAddress'):
                            address_components.append(user.get('streetAddress'))
                        if user.get('city'):
                            address_components.append(user.get('city'))
                        if user.get('state'):
                            address_components.append(user.get('state'))
                        if user.get('postalCode'):
                            address_components.append(user.get('postalCode'))
                        if user.get('country'):
                            address_components.append(user.get('country'))

                        full_address = ', '.join(address_components) if address_components else ''

                        email_value = primary_email or user_principal_name or ''

                        employee = {
                            'id': user.get('id'),
                            'name': display_name or 'Unknown',
                            'title': user.get('jobTitle') or 'No Title',
                            'department': department_val or 'No Department',
                            'email': email_value,
                            'phone': user.get('mobilePhone') or '',
                            'businessPhone': business_phone,
                            'location': user.get('officeLocation') or '',
                            'officeLocation': user.get('officeLocation') or '',
                            'city': user.get('city') or '',
                            'state': user.get('state') or '',
                            'country': user.get('country') or '',
                            'fullAddress': full_address,
                            'managerId': user.get('manager', {}).get('id') if user.get('manager') else None,
                            'employeeHireDate': hire_date_str,
                            'hireDate': hire_date.isoformat() if hire_date else None,
                            'isNewEmployee': is_new,
                            'photoUrl': f'/api/photo/{user.get("id")}',
                            'userPrincipalName': user_principal_name,
                            'children': [],
                            'accountEnabled': user.get('accountEnabled', True),
                            'userType': user.get('userType') or '',
                            'usageLocation': user.get('usageLocation') or ''
                        }
                        employees.append(employee)
            
            users_url = data.get('@odata.nextLink')
        except requests.exceptions.RequestException as e:
            fetch_failed = True
            logger.error(f"Error fetching employees: {e}")
            status_code = getattr(getattr(e, 'response', None), 'status_code', None)
            if status_code == 401:
                logger.error("Authentication failed. Please check your credentials.")
            elif status_code == 403:
                logger.error("Permission denied. Ensure User.Read.All permission is granted.")
            break
        except Exception as e:
            fetch_failed = True
            logger.error(f"Unexpected error: {e}")
            break
    
    logger.info(
        "Fetched %s employees from Graph API (filtered total %s, with licenses %s)",
        len(employees),
        len(filtered_users),
        len(filtered_with_license)
    )

    if fetch_failed or not employees:
        fallback_employees, fallback_filtered_with_license, fallback_filtered_users = _load_fetch_all_employees_fallback()
        if fallback_employees:
            logger.warning(
                "Using cached employee fallback after Graph fetch %s (%s records)",
                "failure" if fetch_failed else "returning no data",
                len(fallback_employees)
            )
            employees = fallback_employees
            if fallback_filtered_with_license:
                filtered_with_license = fallback_filtered_with_license
            if fallback_filtered_users:
                filtered_users = fallback_filtered_users
        elif fetch_failed:
            logger.warning("Graph fetch failed and no cached employee data is available")

    return employees, filtered_with_license, filtered_users


def _load_fetch_all_employees_fallback():
    cached_employees = load_cached_employees() or []

    def _load_cached_list(path, description):
        if not os.path.exists(path):
            logger.debug(f"No cached {description} found at {path}")
            return []
        try:
            with open(path, 'r') as cache_file:
                data = json.load(cache_file)
        except Exception as error:
            logger.error(f"Failed to load cached {description} from {path}: {error}")
            return []

        if not isinstance(data, list):
            logger.warning(f"Cached {description} at {path} is not a list; ignoring contents")
            return []

        return data

    cached_filtered_with_license = _load_cached_list(FILTERED_LICENSE_FILE, 'filtered licensed users')
    cached_filtered_users = _load_cached_list(FILTERED_USERS_FILE, 'filtered users')

    return cached_employees, cached_filtered_with_license, cached_filtered_users


def fetch_subscribed_sku_map(token):
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    sku_map = {}
    skus_url = f'{GRAPH_API_ENDPOINT}/subscribedSkus?$select=skuId,skuPartNumber'

    try:
        while skus_url:
            response = requests.get(skus_url, headers=headers, timeout=10)
            response.raise_for_status()
            data = response.json()

            for sku in data.get('value', []):
                sku_id = sku.get('skuId')
                if not sku_id:
                    continue
                key = str(sku_id).lower()
                sku_map[key] = sku.get('skuPartNumber') or str(sku_id)

            skus_url = data.get('@odata.nextLink')
    except requests.exceptions.RequestException as error:
        logger.warning(f"Failed to load subscribed SKUs: {error}")
    except Exception as error:
        logger.warning(f"Unexpected error loading subscribed SKUs: {error}")

    return sku_map


def collect_last_login_records(*, token=None):
    """Return cached-friendly last sign-in details for all users."""

    token = token or get_access_token()
    if not token:
        logger.error("Failed to get access token for last sign-in report")
        return []

    sku_map = fetch_subscribed_sku_map(token)

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
        'ConsistencyLevel': 'eventual'
    }

    fields = (
        'id,displayName,jobTitle,department,mail,userPrincipalName,'
        'signInActivity,accountEnabled,userType,assignedLicenses'
    )
    users_url = f"{GRAPH_API_BETA_ENDPOINT}/users?$select={fields}&$top=999"

    now_utc = datetime.now(timezone.utc)
    records = []

    def _format_datetime(dt):
        if not dt:
            return None
        if not dt.tzinfo:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).isoformat()

    def _map_licenses(license_entries):
        license_entries = license_entries or []
        if not license_entries:
            return [], []

        sku_ids = []
        labels = []
        seen_labels = set()

        for entry in license_entries:
            sku_id = entry.get('skuId')
            if not sku_id:
                continue
            sku_ids.append(str(sku_id))
            lookup_key = str(sku_id).lower()
            friendly = sku_map.get(lookup_key) or sku_map.get(lookup_key.upper()) or str(sku_id)
            normalized = friendly.lower()
            if normalized not in seen_labels:
                seen_labels.add(normalized)
                labels.append(friendly)

        labels.sort(key=lambda item: item.lower())
        return sku_ids, labels

    while users_url:
        try:
            response = requests.get(users_url, headers=headers, timeout=20)
        except requests.exceptions.RequestException as error:
            logger.error(f"Failed to fetch sign-in activity: {error}")
            break

        if response.status_code == 429:
            retry_after = response.headers.get('Retry-After')
            delay = 5
            try:
                parsed = int(retry_after)
                delay = max(parsed, delay)
            except Exception:
                pass
            logger.warning(f"Graph throttled sign-in activity request; retrying in {delay} seconds")
            time.sleep(delay)
            continue

        try:
            response.raise_for_status()
        except requests.exceptions.HTTPError as error:
            status_code = getattr(error.response, 'status_code', None)
            logger.error(f"Graph error fetching sign-in activity (status {status_code}): {error}")
            break

        payload = response.json()

        for user in payload.get('value', []):
            sign_in = user.get('signInActivity') or {}

            last_combined = parse_graph_datetime(sign_in.get('lastSignInDateTime'))
            last_interactive = parse_graph_datetime(sign_in.get('lastInteractiveSignInDateTime'))
            last_non_interactive = parse_graph_datetime(sign_in.get('lastNonInteractiveSignInDateTime'))

            observed_dates = [dt for dt in (last_combined, last_interactive, last_non_interactive) if dt]
            most_recent = max(observed_dates) if observed_dates else None

            sku_ids, license_labels = _map_licenses(user.get('assignedLicenses'))

            record = {
                'id': user.get('id'),
                'name': user.get('displayName') or 'Unknown',
                'title': user.get('jobTitle') or 'No Title',
                'department': user.get('department') or 'No Department',
                'email': user.get('mail') or user.get('userPrincipalName') or '',
                'accountEnabled': user.get('accountEnabled', True),
                'userType': (user.get('userType') or '').lower(),
                'licenseCount': len(sku_ids),
                'licenseSkus': license_labels,
                'licenseSkuIds': sku_ids,
                'lastActivityDate': _format_datetime(most_recent),
                'daysSinceLastActivity': int((now_utc - most_recent).days) if most_recent else None,
                'lastInteractiveSignIn': _format_datetime(last_interactive),
                'daysSinceInteractiveSignIn': int((now_utc - last_interactive).days) if last_interactive else None,
                'lastNonInteractiveSignIn': _format_datetime(last_non_interactive),
                'daysSinceNonInteractiveSignIn': int((now_utc - last_non_interactive).days) if last_non_interactive else None,
                'neverSignedIn': not observed_dates,
            }

            records.append(record)

        users_url = payload.get('@odata.nextLink')

    logger.info("Collected %s last sign-in records", len(records))
    return records


def _collect_disabled_users(*, token=None, settings=None):
    """Return disabled user records with optional license metadata."""

    token = token or get_access_token()
    if not token:
        logger.error("Failed to get access token for disabled user reports")
        return []

    sku_map = fetch_subscribed_sku_map(token)

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    select_fields = (
        'id,displayName,jobTitle,department,mail,userPrincipalName,mobilePhone,'
        'businessPhones,officeLocation,city,state,country,usageLocation,streetAddress,'
        'postalCode,employeeHireDate,employeeLeaveDateTime,accountEnabled,userType,assignedLicenses'
    )

    users_url = f'{GRAPH_API_ENDPOINT}/users?$select={select_fields}&$filter=accountEnabled eq false'
    records = []

    while users_url:
        try:
            response = requests.get(users_url, headers=headers, timeout=15)
            response.raise_for_status()
            data = response.json()

            for user in data.get('value', []):
                display_name = user.get('displayName') or ''
                primary_email = user.get('mail') or ''
                user_principal_name = user.get('userPrincipalName') or ''
                job_title_val = user.get('jobTitle') or ''
                department_val = user.get('department') or 'No Department'

                business_phones = user.get('businessPhones') or []
                if isinstance(business_phones, list):
                    business_phone = next((phone for phone in business_phones if phone), '')
                else:
                    business_phone = business_phones or ''

                assigned_licenses = user.get('assignedLicenses') or []
                license_sku_ids = []
                license_labels = []
                for license_entry in assigned_licenses:
                    sku_id = license_entry.get('skuId')
                    if not sku_id:
                        continue
                    sku_key = str(sku_id).lower()
                    license_sku_ids.append(str(sku_id))
                    friendly_name = (
                        sku_map.get(sku_key)
                        or sku_map.get(sku_key.upper())
                        or str(sku_id)
                    )
                    license_labels.append(friendly_name)

                license_labels = sorted(set(license_labels), key=lambda item: item.lower()) if license_labels else []

                disabled_at = parse_graph_datetime(user.get('employeeLeaveDateTime'))
                disabled_iso = datetime_to_iso(disabled_at) if disabled_at else None

                hire_date = parse_graph_datetime(user.get('employeeHireDate'))

                record = {
                    'id': user.get('id'),
                    'name': display_name or 'Unknown',
                    'title': job_title_val or 'No Title',
                    'department': department_val,
                    'email': primary_email or user_principal_name or '',
                    'userPrincipalName': user_principal_name,
                    'phone': user.get('mobilePhone') or '',
                    'businessPhone': business_phone,
                    'location': user.get('officeLocation') or '',
                    'city': user.get('city') or '',
                    'state': user.get('state') or '',
                    'country': user.get('country') or '',
                    'usageLocation': user.get('usageLocation') or '',
                    'accountEnabled': user.get('accountEnabled', True),
                    'userType': (user.get('userType') or '').lower(),
                    'licenseCount': len(license_sku_ids),
                    'licenseSkus': license_labels,
                    'licenseSkuIds': license_sku_ids,
                    'hireDate': datetime_to_iso(hire_date) if hire_date else None,
                    'disabledDate': disabled_iso,
                    'disabledDays': calculate_days_since(disabled_at),
                }

                records.append(record)

            users_url = data.get('@odata.nextLink')
        except requests.exceptions.RequestException as error:
            logger.error(f"Error fetching disabled users: {error}")
            break
        except Exception as error:
            logger.error(f"Unexpected error while collecting disabled user data: {error}")
            break

    logger.info(f"Collected {len(records)} disabled users")
    return records


def collect_disabled_users(*, token=None, settings=None, previous_records=None):
    raw_records = _collect_disabled_users(token=token, settings=settings)

    previous_map = {}
    if previous_records:
        for entry in previous_records:
            entry_id = entry.get('id')
            if entry_id:
                previous_map[entry_id] = entry

    now_iso = datetime_to_iso(datetime.now(timezone.utc))

    for record in raw_records:
        record_id = record.get('id')
        existing = previous_map.get(record_id) if record_id else None

        observed_source = record.get('disabledDate')
        existing_observed = None

        if existing:
            existing_observed = (
                existing.get('firstSeenDisabledAt')
                or existing.get('disabledDate')
            )

        if observed_source:
            first_seen = observed_source
        elif existing_observed:
            first_seen = existing_observed
        else:
            first_seen = now_iso

        record['firstSeenDisabledAt'] = first_seen

        if not record.get('disabledDate'):
            record['disabledDate'] = first_seen

        record['disabledDays'] = calculate_days_since(first_seen)

    return raw_records


def collect_disabled_licensed_users(*, token=None, settings=None, previous_records=None):
    raw_records = collect_disabled_users(
        token=token,
        settings=settings,
        previous_records=previous_records
    )
    licensed_records = [record for record in raw_records if (record.get('licenseCount') or 0) > 0]
    logger.info(f"Filtered {len(licensed_records)} disabled users with active licenses")
    return licensed_records

def build_org_hierarchy(employees, *, top_user_email_override=None, settings=None):
    if not employees:
        return None
    
    if settings is None:
        settings = load_settings()

    settings_top_user = (settings.get('topUserEmail') or '').strip()
    env_top_user = (TOP_LEVEL_USER_EMAIL or '').strip()

    if top_user_email_override is not None:
        chosen_top_user = (top_user_email_override or '').strip()
    elif env_top_user:
        chosen_top_user = env_top_user
    else:
        chosen_top_user = settings_top_user

    top_user_email = (chosen_top_user or '').strip() or None

    # Debug logging
    logger.info(f"Settings topUserEmail: '{settings_top_user}'")
    logger.info(f"Environment TOP_LEVEL_USER_EMAIL: '{env_top_user}'")
    if top_user_email_override is not None:
        logger.info(f"Session override topUserEmail: '{top_user_email_override}'")
    logger.info(f"Final top_user_email: '{top_user_email}'")
    logger.info(f"TOP_LEVEL_USER_ID: '{TOP_LEVEL_USER_ID}'")
    
    emp_dict = {emp['id']: emp.copy() for emp in employees}
    
    for emp_id in emp_dict:
        if 'children' not in emp_dict[emp_id]:
            emp_dict[emp_id]['children'] = []
    
    # First, check if a specific top-level user is configured
    # Prioritize settings file email over environment variables
    root = None
    if top_user_email:
        logger.info(f"Searching for user with email: '{top_user_email}' among {len(employees)} employees")
        for emp in employees:
            if emp.get('email') == top_user_email:
                root = emp_dict[emp['id']]
                logger.info(f"Found and using configured top-level user by email: {root['name']} ({root.get('email')})")
                break
        else:
            logger.warning(f"Could not find user with email '{top_user_email}' in employee list")
    
    # Fallback to environment variable ID if no email-based selection was made
    if not root and TOP_LEVEL_USER_ID and TOP_LEVEL_USER_ID in emp_dict:
        root = emp_dict[TOP_LEVEL_USER_ID]
        logger.info(f"Using fallback environment top-level user by ID: {root['name']}")
    
    if root:
        # If a specific root is configured, build hierarchy with that person at the top
        # Clear any existing manager relationship for the root user
        root['managerId'] = None
        
        # Build the hierarchy normally but ensure the selected root has no manager
        for emp in employees:
            emp_copy = emp_dict[emp['id']]
            if emp_copy['id'] == root['id']:
                continue  # Skip the root user in hierarchy building
                
            if emp['managerId'] and emp['managerId'] in emp_dict:
                manager = emp_dict[emp['managerId']]
                if emp_copy not in manager['children']:
                    manager['children'].append(emp_copy)
        
        # Remove the selected root from anyone's children list (in case they were someone's subordinate)
        for emp_id, emp in emp_dict.items():
            emp['children'] = [child for child in emp['children'] if child['id'] != root['id']]
        
        return root
    else:
        # Auto-detect root using existing logic
        root_candidates = []
        
        # Build normal manager-employee relationships
        for emp in employees:
            emp_copy = emp_dict[emp['id']]
            if emp['managerId'] and emp['managerId'] in emp_dict:
                manager = emp_dict[emp['managerId']]
                if emp_copy not in manager['children']:
                    manager['children'].append(emp_copy)
            else:
                if not emp['managerId'] and emp_copy not in root_candidates:
                    root_candidates.append(emp_copy)
        
        # Auto-detect root
        if root_candidates:
            ceo_keywords = ['chief executive', 'ceo', 'president', 'chair', 'director', 'head']
            for candidate in root_candidates:
                title_lower = (candidate.get('title') or '').lower()
                if any(keyword in title_lower for keyword in ceo_keywords):
                    root = candidate
                    logger.info(f"Auto-detected top-level user: {root['name']} - {root.get('title')}")
                    break
            
            if not root and root_candidates:
                root = root_candidates[0]
                logger.info(f"Using first root candidate as top-level: {root['name']}")
        else:
            max_reports = 0
            for emp_id, emp in emp_dict.items():
                if len(emp['children']) > max_reports:
                    max_reports = len(emp['children'])
                    root = emp
            
            if root:
                logger.info(f"Using person with most reports as top-level: {root['name']} ({max_reports} reports)")
        
        if not root and employees:
            root = emp_dict[employees[0]['id']]
            logger.info(f"Using first employee as root: {root['name']}")
        
        return root


def collect_missing_manager_records(employees, hierarchy_root=None, settings=None, top_user_email_override=None):
    if not employees:
        return []

    employee_index = {emp['id']: emp for emp in employees if emp.get('id')}
    visited = set()

    def traverse(node):
        node_id = node.get('id')
        if not node_id or node_id in visited:
            return
        visited.add(node_id)
        for child in node.get('children', []):
            traverse(child)

    if hierarchy_root:
        traverse(hierarchy_root)

    root_ids = set()
    top_user_email = None

    if hierarchy_root and hierarchy_root.get('id'):
        root_ids.add(hierarchy_root['id'])

    if settings is None:
        settings = load_settings()

    if top_user_email_override is not None:
        top_user_email = (top_user_email_override or '').strip().lower() or None
    elif settings:
        top_user_email = (settings.get('topUserEmail') or '').strip().lower() or None
    elif TOP_LEVEL_USER_EMAIL:
        top_user_email = (TOP_LEVEL_USER_EMAIL or '').strip().lower() or None

    missing_records = []

    for emp in employees:
        emp_id = emp.get('id')
        manager_id = emp.get('managerId')
        manager_name = ''
        reason = None

        if emp_id and emp_id in root_ids:
            continue

        if top_user_email:
            email = (emp.get('email') or '').strip().lower()
            if email and email == top_user_email:
                continue

        if manager_id and manager_id in employee_index:
            manager_name = employee_index[manager_id].get('name') or ''

        if not manager_id:
            reason = 'no_manager'
        elif manager_id not in employee_index:
            reason = 'manager_not_found'
        elif emp_id not in visited:
            reason = 'detached'

        if reason:
            missing_records.append({
                'id': emp_id,
                'name': emp.get('name'),
                'title': emp.get('title'),
                'department': emp.get('department'),
                'email': emp.get('email'),
                'phone': emp.get('phone'),
                'businessPhone': emp.get('businessPhone'),
                'location': emp.get('location') or emp.get('officeLocation') or '',
                'managerName': manager_name,
                'reason': reason
            })

    missing_records.sort(key=lambda item: (item.get('department') or '', item.get('name') or ''))
    return missing_records


def update_employee_data():
    try:
        # Ensure data directory exists and is writable
        if not os.path.exists(DATA_DIR):
            os.makedirs(DATA_DIR, exist_ok=True)
            logger.info(f"Created data directory: {DATA_DIR}")

        # Test if we can write to the data directory
        test_file = os.path.join(DATA_DIR, 'test_write.tmp')
        try:
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
        except Exception as e:
            logger.error(f"Cannot write to data directory {DATA_DIR}: {e}")
            return

        logger.info(f"[{datetime.now()}] Starting employee data update...")

        token = get_access_token()
        if not token:
            logger.error("Unable to refresh employee data because access token retrieval failed")
            return

        settings = load_settings()
        months_threshold = settings.get('newEmployeeMonths', 3)

        employees, filtered_with_license, filtered_users = fetch_all_employees(token=token)

        if employees:
            ignored_employee_set = parse_ignored_employees(settings)
            ignored_department_set = parse_ignored_departments(settings)

            if ignored_employee_set:
                before = len(employees)
                employees = [
                    emp for emp in employees
                    if not employee_is_ignored(
                        emp.get('name'),
                        emp.get('email'),
                        emp.get('userPrincipalName'),
                        ignored_employee_set
                    )
                ]
                if before != len(employees):
                    logger.info(f"Filtered ignored employees; {before}->{len(employees)} remaining")

            if ignored_department_set:
                before = len(employees)
                employees = [
                    emp for emp in employees
                    if not department_is_ignored(emp.get('department'), ignored_department_set)
                ]
                logger.info(
                    f"Filtered ignored departments {sorted(list(ignored_department_set))}; {before}->{len(employees)} employees"
                )

            try:
                with open(EMPLOYEE_LIST_FILE, 'w') as employee_cache:
                    json.dump(employees, employee_cache, indent=2)
                logger.info(f"Cached {len(employees)} employees for session-specific hierarchy builds")
            except Exception as cache_error:
                logger.error(f"Failed to write employee cache: {cache_error}")

            hierarchy = build_org_hierarchy(employees, settings=settings)
            missing_records = collect_missing_manager_records(employees, hierarchy, settings)

            if hierarchy:
                def update_new_status(node):
                    if node.get('hireDate'):
                        try:
                            hire_date = datetime.fromisoformat(node['hireDate'])
                            if hire_date.tzinfo:
                                cutoff_date = datetime.now(hire_date.tzinfo) - timedelta(days=months_threshold * 30)
                            else:
                                cutoff_date = datetime.now() - timedelta(days=months_threshold * 30)
                            node['isNewEmployee'] = hire_date > cutoff_date
                        except Exception:
                            node['isNewEmployee'] = False
                    else:
                        node['isNewEmployee'] = False

                    for child in node.get('children', []) or []:
                        update_new_status(child)

                update_new_status(hierarchy)

                with open(DATA_FILE, 'w') as f:
                    json.dump(hierarchy, f, indent=2)
                logger.info(f"[{datetime.now()}] Successfully updated employee data. Total employees: {len(employees)}")

                try:
                    with open(MISSING_MANAGER_FILE, 'w') as report_file:
                        json.dump(missing_records, report_file, indent=2)
                    logger.info(f"Updated missing manager report cache with {len(missing_records)} records")
                except Exception as report_error:
                    logger.error(f"Failed to write missing manager report cache: {report_error}")
            else:
                logger.error(f"[{datetime.now()}] Could not build hierarchy from employee data")

            try:
                recently_hired_records = collect_recently_hired_employees(employees, days=365)
                with open(RECENTLY_HIRED_FILE, 'w') as report_file:
                    json.dump(recently_hired_records, report_file, indent=2)
                logger.info(
                    f"Updated recently hired employees report cache with {len(recently_hired_records)} records"
                )
            except Exception as report_error:
                logger.error(f"Failed to write recently hired employees report cache: {report_error}")
        else:
            logger.error(f"[{datetime.now()}] No employees fetched from Graph API")

        try:
            filtered_user_records = filtered_users or []
            with open(FILTERED_USERS_FILE, 'w') as report_file:
                json.dump(filtered_user_records, report_file, indent=2)
            logger.info(
                f"Updated filtered users report cache with {len(filtered_user_records)} records"
            )
        except Exception as report_error:
            logger.error(f"Failed to write filtered users report cache: {report_error}")

        try:
            filtered_license_records = filtered_with_license or []
            with open(FILTERED_LICENSE_FILE, 'w') as report_file:
                json.dump(filtered_license_records, report_file, indent=2)
            logger.info(
                f"Updated filtered licensed users report cache with {len(filtered_license_records)} records"
            )
        except Exception as report_error:
            logger.error(f"Failed to write filtered licensed users report cache: {report_error}")

        try:
            last_login_records = collect_last_login_records(token=token)
            with open(LAST_LOGIN_FILE, 'w') as report_file:
                json.dump(last_login_records, report_file, indent=2)
            logger.info(
                f"Updated last sign-in report cache with {len(last_login_records)} records"
            )
        except Exception as report_error:
            logger.error(f"Failed to write last sign-in report cache: {report_error}")

        try:
            existing_disabled_records = []
            if os.path.exists(DISABLED_USERS_FILE):
                try:
                    with open(DISABLED_USERS_FILE, 'r') as previous_file:
                        data = json.load(previous_file)
                        if isinstance(data, list):
                            existing_disabled_records = data
                except Exception as previous_error:
                    logger.warning(f"Unable to load existing disabled users cache: {previous_error}")

            disabled_user_records = collect_disabled_users(
                token=token,
                settings=settings,
                previous_records=existing_disabled_records
            ) or []
            with open(DISABLED_USERS_FILE, 'w') as report_file:
                json.dump(disabled_user_records, report_file, indent=2)
            logger.info(
                f"Updated disabled users report cache with {len(disabled_user_records)} records"
            )
        except Exception as report_error:
            logger.error(f"Failed to write disabled users report cache: {report_error}")

        try:
            disabled_license_records = [
                record for record in disabled_user_records if (record.get('licenseCount') or 0) > 0
            ]

            with open(DISABLED_LICENSE_FILE, 'w') as report_file:
                json.dump(disabled_license_records, report_file, indent=2)
            logger.info(
                f"Updated disabled licensed users report cache with {len(disabled_license_records)} records"
            )
        except Exception as report_error:
            logger.error(f"Failed to write disabled licensed users report cache: {report_error}")

        try:
            recently_disabled_records = collect_recently_disabled_employees(disabled_user_records, days=365)
            with open(RECENTLY_DISABLED_FILE, 'w') as report_file:
                json.dump(recently_disabled_records, report_file, indent=2)
            logger.info(
                f"Updated recently disabled employees report cache with {len(recently_disabled_records)} records"
            )
        except Exception as report_error:
            logger.error(f"Failed to write recently disabled employees report cache: {report_error}")
    except Exception as e:
        logger.error(f"[{datetime.now()}] Error updating employee data: {e}")


def load_cached_employees():
    if os.path.exists(EMPLOYEE_LIST_FILE):
        try:
            with open(EMPLOYEE_LIST_FILE, 'r') as cache_file:
                return json.load(cache_file)
        except Exception as e:
            logger.error(f"Failed to read employee cache {EMPLOYEE_LIST_FILE}: {e}")
    return None


def flatten_hierarchy_to_employee_list(root_node):
    employees = []

    def _walk(node):
        if not isinstance(node, dict):
            return

        entry = {k: v for k, v in node.items() if k != 'children'}
        entry['children'] = []
        employees.append(entry)

        for child in node.get('children', []) or []:
            _walk(child)

    if root_node:
        _walk(root_node)

    return employees


def get_employee_list_for_metadata():
    employees = load_cached_employees()
    if employees:
        return employees

    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r') as data_file:
                hierarchy = json.load(data_file)
            if hierarchy:
                return flatten_hierarchy_to_employee_list(hierarchy)
        except Exception as error:
            logger.error(f"Failed to read hierarchy for metadata: {error}")

    return []


def collect_unique_field_values(employees, field_name):
    unique = {}
    for employee in employees or []:
        value = (employee.get(field_name) or '').strip()
        if not value:
            continue
        key = value.lower()
        if key not in unique:
            unique[key] = value

    return sorted(unique.values(), key=lambda item: item.lower())


def collect_employee_option_labels(employees):
    options = {}
    for employee in employees or []:
        name = (employee.get('name') or '').strip()
        email = (employee.get('email') or '').strip()
        user_principal_name = (employee.get('userPrincipalName') or '').strip()

        contact = email or user_principal_name

        if not name and not contact:
            continue

        if name and contact:
            label = f"{name} <{contact}>"
        else:
            label = name or contact

        primary_key = normalize_filter_value(contact) or normalize_filter_value(name) or normalize_filter_value(label)
        if not primary_key:
            continue

        if primary_key not in options:
            options[primary_key] = label

    return sorted(options.values(), key=lambda item: item.lower())

def schedule_updates():
    global scheduler_running
    
    settings = load_settings()
    
    if os.environ.get('RUN_INITIAL_UPDATE', 'true').lower() == 'true':
        logger.info(f"[{datetime.now()}] Running initial employee data update on startup...")
        update_employee_data()
    
    if settings.get('autoUpdateEnabled', True):
        update_time = settings.get('updateTime', '20:00')
        schedule.every().day.at(update_time).do(update_employee_data)
        logger.info(f"Scheduled daily updates at {update_time}")
    
    while scheduler_running:
        schedule.run_pending()
        time.sleep(60)

def start_scheduler():
    global scheduler_running
    with scheduler_lock:
        if not scheduler_running:
            scheduler_running = True
            scheduler_thread = threading.Thread(target=schedule_updates, daemon=True)
            scheduler_thread.start()
            logger.info("Scheduler started")

def stop_scheduler():
    global scheduler_running
    with scheduler_lock:
        scheduler_running = False
        logger.info("Scheduler stopped")

def restart_scheduler():
    """Restart scheduler with new settings"""
    stop_scheduler()
    time.sleep(2)
    schedule.clear()
    start_scheduler()


if hasattr(app, 'before_serving'):

    @app.before_serving
    def _start_scheduler_when_ready():
        start_scheduler()


    @app.after_serving
    def _stop_scheduler_on_shutdown():
        stop_scheduler()

elif hasattr(app, 'before_request'):

    @app.before_request
    def _ensure_scheduler_started():
        if not scheduler_running:
            start_scheduler()


atexit.register(stop_scheduler)

def get_template(template_name):
    """Load HTML template from file"""
    possible_paths = [
        f'templates/{template_name}',
        template_name,
        os.path.join(os.path.dirname(__file__), 'templates', template_name),
        os.path.join(os.path.dirname(__file__), template_name)
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    # Only log template loading in debug mode to reduce log spam
                    logger.debug(f"Loading template from: {path}")
                    return f.read()
            except Exception as e:
                logger.error(f"Error reading {path}: {e}")
    
    logger.error(f"{template_name} not found in any expected location")
    return f"<h1>Error: {template_name} not found</h1>"

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    next_page = sanitize_next_path(request.args.get('next', ''))

    if request.method == 'POST':
        try:
            payload = request.get_json(silent=True)
            password = None

            if isinstance(payload, dict):
                password = payload.get('password')

            if password is None:
                password = request.form.get('password')

            password = password or ''

            logger.info(f"Login attempt from {get_remote_address()} - IP address authentication check")

            if password == ADMIN_PASSWORD:
                session['authenticated'] = True
                session['username'] = 'admin'
                logger.info("Successful login")
                return jsonify({
                    'success': True,
                    'next': next_page or ''
                })

            logger.warning("Failed login attempt - password mismatch")
            return jsonify({'error': 'Invalid password'}), 401
        except Exception as e:
            logger.error(f"Login error: {e}")
            return jsonify({'error': 'Login failed'}), 500

    settings = load_settings()
    chart_title = (settings.get('chartTitle') or '').strip() or 'DB AutoOrgChart'
    favicon_path = settings.get('faviconPath', '/favicon.ico')

    return render_template(
        'login.html',
        chart_title=chart_title,
        next_page=next_page,
        favicon_path=favicon_path
    )

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/')
def index():
    template_content = get_template('index.html')
    settings = load_settings()
    favicon_path = settings.get('faviconPath', '/favicon.ico')
    
    # Inject favicon link into template
    favicon_link = f'<link rel="icon" type="image/x-icon" href="{favicon_path}">'
    template_content = template_content.replace('</head>', f'    {favicon_link}\n</head>')
    
    return render_template_string(template_content)

@app.route('/configure')
def configure():
    if not session.get('authenticated'):
        # Redirect to login preserving the intended destination
        desired_path = sanitize_next_path(request.path)
        params = {'next': desired_path} if desired_path else {}
        return redirect(url_for('login', **params))
    
    template_content = get_template('configure.html')
    settings = load_settings()
    favicon_path = settings.get('faviconPath', '/favicon.ico')
    chart_title = (settings.get('chartTitle') or '').strip() or 'DB AutoOrgChart'
    
    # Inject favicon link into template
    favicon_link = f'<link rel="icon" type="image/x-icon" href="{favicon_path}">'
    template_content = template_content.replace('</head>', f'    {favicon_link}\n</head>')
    
    return render_template_string(
        template_content,
        chart_title=chart_title,
        _=translate_placeholder
    )


@app.route('/reports')
@login_required
def reports():
    template_content = get_template('reports.html')
    settings = load_settings()
    favicon_path = settings.get('faviconPath', '/favicon.ico')

    favicon_link = f'<link rel="icon" type="image/x-icon" href="{favicon_path}">'
    template_content = template_content.replace('</head>', f'    {favicon_link}\n</head>')

    return render_template_string(template_content)

@app.route('/static/icon_custom_<string:file_hash>.png')
def serve_custom_logo(file_hash):
    """Serve custom logo files from the data directory"""
    # Validate file_hash to prevent directory traversal
    if not file_hash.isalnum() or len(file_hash) != 8:
        return "Invalid logo identifier", 400
    
    custom_logo = os.path.join(DATA_DIR, f'icon_custom_{file_hash}.png')
    
    if os.path.exists(custom_logo) and os.path.isfile(custom_logo):
        return send_file(custom_logo, 
                       mimetype='image/png',
                       as_attachment=False,
                       max_age=3600)  # Cache for 1 hour
    else:
        return "Logo not found", 404

@app.route('/static/favicon_custom_<file_hash>.<ext>')
def serve_custom_favicon(file_hash, ext):
    """Serve custom favicon files from the data directory"""
    # Validate file_hash to prevent directory traversal
    if not file_hash.isalnum() or len(file_hash) != 12:
        return "Invalid favicon identifier", 400
    
    # Validate extension
    if ext not in ['ico', 'png']:
        return "Invalid favicon format", 400
    
    custom_favicon = os.path.join(DATA_DIR, f'favicon_custom_{file_hash}.{ext}')
    
    if os.path.exists(custom_favicon) and os.path.isfile(custom_favicon):
        mimetype = 'image/x-icon' if ext == 'ico' else 'image/png'
        return send_file(custom_favicon, 
                       mimetype=mimetype,
                       as_attachment=False,
                       max_age=3600)  # Cache for 1 hour
    else:
        return "Favicon not found", 404

@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve regular static files"""
    return send_from_directory('static', filename)

@app.route('/api/photo/<user_id>')
@limiter.limit("500 per hour")  # Higher limit for photo endpoint due to org chart loading
def get_employee_photo(user_id):
    """Serve employee photo from Microsoft Graph API with caching"""
    try:
        # Create photos cache directory
        photos_dir = os.path.join(DATA_DIR, 'photos')
        if not os.path.exists(photos_dir):
            os.makedirs(photos_dir, exist_ok=True)
        
        # Check if photo is cached
        photo_file = os.path.join(photos_dir, f"{user_id}.jpg")
        
        if os.path.exists(photo_file):
            # Check if cache is less than 1 day old
            if time.time() - os.path.getmtime(photo_file) < 86400:  # 24 hours
                response = send_file(photo_file, mimetype='image/jpeg')
                # Add cache headers to prevent browser caching issues
                response.headers['Cache-Control'] = 'public, max-age=3600'
                response.headers['Last-Modified'] = datetime.fromtimestamp(os.path.getmtime(photo_file)).strftime('%a, %d %b %Y %H:%M:%S GMT')
                return response
        
        # Fetch fresh photo from Graph API
        token = get_access_token()
        if token:
            photo_data = fetch_employee_photo(user_id, token)
            if photo_data:
                # Save to cache
                with open(photo_file, 'wb') as f:
                    f.write(photo_data)
                
                # Serve the photo
                response = send_file(
                    BytesIO(photo_data),
                    mimetype='image/jpeg',
                    as_attachment=False
                )
                response.headers['Cache-Control'] = 'public, max-age=3600'
                return response
        
        # Fallback to default user icon
        logger.debug(f"No photo available for user {user_id}, using fallback")
        return send_from_directory('static', 'usericon.png')
        
    except Exception as e:
        logger.error(f"Error serving photo for user {user_id}: {e}")
        return send_from_directory('static', 'usericon.png')

@app.route('/api/employees')
def get_employees():
    try:
        logger.info("API request for /api/employees received")
        if not os.path.exists(DATA_FILE):
            logger.info("Data file does not exist, attempting to create it...")
            update_employee_data()
            
        # Double check the file exists after update attempt
        if not os.path.exists(DATA_FILE):
            logger.error(f"Could not create data file {DATA_FILE}")
            return jsonify({'error': 'No employee data available. Please check configuration.'}), 500
        
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)

        settings = load_settings()
        months_threshold = settings.get('newEmployeeMonths', 3)

        session_override_present = 'topUserEmail' in session
        session_top_user = (session.get('topUserEmail') or '').strip() if session_override_present else None
        env_top_user = (TOP_LEVEL_USER_EMAIL or '').strip()

        requested_top_user = None
        override_reason = None

        if session_override_present:
            requested_top_user = session_top_user
            override_reason = 'session override'
        elif env_top_user:
            current_root_email = (data or {}).get('email') or ''
            if current_root_email.strip().lower() != env_top_user.lower():
                requested_top_user = env_top_user
                override_reason = 'environment default enforcement'

        if requested_top_user is not None:
            employees = load_cached_employees()
            if not employees and data:
                employees = flatten_hierarchy_to_employee_list(data)
            if not employees:
                logger.info("Employee cache unavailable; fetching employees from Graph API for top user override")
                employees, _, _ = fetch_all_employees()
                if employees:
                    try:
                        with open(EMPLOYEE_LIST_FILE, 'w') as cache_file:
                            json.dump(employees, cache_file, indent=2)
                    except Exception as cache_error:
                        logger.error(f"Failed to refresh employee cache: {cache_error}")

            if employees:
                override_hierarchy = build_org_hierarchy(
                    employees,
                    top_user_email_override=requested_top_user,
                    settings=settings
                )
                if override_hierarchy:
                    data = override_hierarchy

                    if override_reason == 'environment default enforcement':
                        try:
                            with open(DATA_FILE, 'w') as data_file:
                                json.dump(data, data_file, indent=2)
                            missing_records = collect_missing_manager_records(
                                employees,
                                data,
                                settings,
                                top_user_email_override=requested_top_user
                            )
                            with open(MISSING_MANAGER_FILE, 'w') as report_file:
                                json.dump(missing_records, report_file, indent=2)
                            logger.info("Refreshed global hierarchy cache to align with environment top user")
                        except Exception as cache_error:
                            logger.error(f"Failed to persist environment-aligned hierarchy: {cache_error}")
                else:
                    logger.warning("Failed to build hierarchy with requested top user override; returning cached hierarchy")
            else:
                logger.warning("Unable to locate employee data while applying top user override; returning cached hierarchy")
        
        if data:
            def update_new_status(node):
                if node.get('hireDate'):
                    try:
                        hire_date = datetime.fromisoformat(node['hireDate'])
                        if hire_date.tzinfo:
                            cutoff_date = datetime.now(hire_date.tzinfo) - timedelta(days=months_threshold * 30)
                        else:
                            cutoff_date = datetime.now() - timedelta(days=months_threshold * 30)
                        node['isNewEmployee'] = hire_date > cutoff_date
                    except:
                        node['isNewEmployee'] = False
                else:
                    node['isNewEmployee'] = False
                
                if node.get('children'):
                    for child in node['children']:
                        update_new_status(child)
            
            update_new_status(data)
        
        # Debug logging for root user
        if data and data.get('name'):
            logger.info(f"Returning org chart data with root user: {data['name']} ({data.get('email', 'no email')})")
        
        if not data:
            logger.warning("No hierarchical data available")
            employees, _, _ = fetch_all_employees()
            if employees:
                data = {
                    'id': 'root',
                    'name': 'Organization',
                    'title': 'All Employees',
                    'department': '',
                    'email': '',
                    'phone': '',
                    'businessPhone': '',
                    'location': '',
                    'officeLocation': '',
                    'city': '',
                    'state': '',
                    'country': '',
                    'fullAddress': '',
                    'children': employees
                }
            else:
                data = {
                    'id': 'root',
                    'name': 'No Data',
                    'title': 'Please check configuration',
                    'businessPhone': '',
                    'children': []
                }
        
        return jsonify(data)
    except Exception as e:
        logger.error(f"Error in get_employees: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings', methods=['GET', 'POST'])
def handle_settings():
    if request.method == 'GET':
        # GET is allowed without auth for loading initial settings
        settings = load_settings()
        if 'topUserEmail' in session:
            settings['topUserEmail'] = session.get('topUserEmail') or ''
        return jsonify(settings)
    
    elif request.method == 'POST':
        # POST requires authentication
        if not session.get('authenticated'):
            return jsonify({'error': 'Authentication required'}), 401
            
        try:
            # Simply update settings without validation
            new_settings = request.json
            current_settings = load_settings()
            current_settings.update(new_settings)
            
            if save_settings(current_settings):
                if ('updateTime' in new_settings or 'autoUpdateEnabled' in new_settings):
                    threading.Thread(target=restart_scheduler).start()
                
                return jsonify({'success': True})
            else:
                return jsonify({'error': 'Failed to save settings'}), 500
        except Exception as e:
            logger.error(f"Error updating settings: {e}")
            return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/metadata/options')
@require_auth
def get_metadata_options():
    employees = get_employee_list_for_metadata()
    job_titles = collect_unique_field_values(employees, 'title')
    departments = collect_unique_field_values(employees, 'department')
    employee_options = collect_employee_option_labels(employees)

    return jsonify({
        'jobTitles': job_titles,
        'departments': departments,
        'employees': employee_options
    })

@app.route('/api/set-top-user', methods=['POST'])
@limiter.limit("20 per minute")
def set_top_user():
    """Store the caller's preferred top-level user in their session"""
    try:
        data = request.json or {}
        if 'topUserEmail' not in data:
            return jsonify({'error': 'Missing topUserEmail parameter'}), 400

        requested_email = (data.get('topUserEmail') or '').strip()
        session['topUserEmail'] = requested_email
        session.modified = True

        logger.info(
            f"Stored session-specific top user preference '{requested_email or 'auto-detect'}' for client {request.remote_addr}"
        )

        return jsonify({'success': True, 'topUserEmail': requested_email})
    except Exception as e:
        logger.error(f"Error updating top-level user session preference: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/set-multiline-enabled', methods=['POST'])
@require_auth
@limiter.limit("20 per minute")
def set_multiline_enabled():
    """Public endpoint to toggle multi-line children layout (Compact Teams)."""
    try:
        data = request.json or {}
        if 'multiLineChildrenEnabled' not in data:
            return jsonify({'error': 'Missing multiLineChildrenEnabled parameter'}), 400

        current_settings = load_settings()
        current_settings['multiLineChildrenEnabled'] = bool(data['multiLineChildrenEnabled'])

        if save_settings(current_settings):
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Failed to save settings'}), 500
    except Exception as e:
        logger.error(f"Error updating multi-line setting: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/test-hierarchy/<email>')
@require_auth
@limiter.limit("5 per minute")
def test_hierarchy(email):
    """Test endpoint to check hierarchy building with specific email"""
    try:
        # Temporarily override settings for testing
        import tempfile
        
        # Fetch fresh employees
        employees, _, _ = fetch_all_employees()
        if not employees:
            return jsonify({'error': 'No employees found'}), 404
            
        # Find the user with this email
        target_user = None
        for emp in employees:
            if emp.get('email') == email:
                target_user = emp
                break
                
        if not target_user:
            return jsonify({'error': f'User with email {email} not found'}), 404
            
        # Temporarily set the environment variable
        original_email = os.environ.get('TOP_LEVEL_USER_EMAIL')
        os.environ['TOP_LEVEL_USER_EMAIL'] = email
        
        try:
            # Build hierarchy
            hierarchy = build_org_hierarchy(employees)
        finally:
            # Restore original
            if original_email:
                os.environ['TOP_LEVEL_USER_EMAIL'] = original_email
            else:
                os.environ.pop('TOP_LEVEL_USER_EMAIL', None)
        
        if hierarchy:
            return jsonify({
                'success': True,
                'root_user': {
                    'name': hierarchy.get('name'),
                    'email': hierarchy.get('email'),
                    'title': hierarchy.get('title')
                },
                'test_email': email,
                'target_user': {
                    'name': target_user.get('name'),
                    'email': target_user.get('email'),
                    'title': target_user.get('title')
                },
                'total_employees': len(employees)
            })
        else:
            return jsonify({'error': 'Failed to build hierarchy'}), 500
            
    except Exception as e:
        logger.error(f"Error in test hierarchy: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-logo', methods=['POST'])
@require_auth
@limiter.limit("5 per minute")
def upload_logo():
    try:
        if 'logo' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['logo']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({'error': f'File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)}MB'}), 400
        
        if file and allowed_file(file.filename):
            # Validate it's actually a safe image
            if not validate_image_file(file):
                return jsonify({'error': 'Invalid or corrupted image file'}), 400
            
            # Generate secure filename
            filename = secure_filename(file.filename)
            file_hash = hashlib.md5(file.read()).hexdigest()[:8]
            file.seek(0)  # Reset after reading for hash
            
            custom_logo_path = os.path.join(DATA_DIR, f'icon_custom_{file_hash}.png')
            
            # Check if directory is writable
            if not os.access(DATA_DIR, os.W_OK):
                return jsonify({'error': 'Server configuration error'}), 500

            file.save(custom_logo_path)
            
            # Verify file was saved
            if not os.path.exists(custom_logo_path):
                return jsonify({'error': 'Failed to save file'}), 500
            
            settings = load_settings()
            settings['logoPath'] = f'/static/icon_custom_{file_hash}.png'
            save_settings(settings)
            
            return jsonify({'success': True, 'path': settings['logoPath']})
        else:
            return jsonify({'error': 'Invalid file type. Only PNG, JPG, JPEG allowed'}), 400
    except Exception as e:
        logger.error(f"Error uploading logo: {e}")
        return jsonify({'error': 'Upload failed'}), 500

@app.route('/api/reset-logo', methods=['POST'])
@require_auth
def reset_logo():
    try:
        # Remove any custom logo files from data directory
        import glob
        custom_logos = glob.glob(os.path.join(DATA_DIR, 'icon_custom*.png'))
        for logo_path in custom_logos:
            if os.path.exists(logo_path):
                os.remove(logo_path)
        
        settings = load_settings()
        settings['logoPath'] = '/static/icon.png'
        save_settings(settings)
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error resetting logo: {e}")
        return jsonify({'error': 'Reset failed'}), 500

@app.route('/api/upload-favicon', methods=['POST'])
@require_auth
def upload_favicon():
    try:
        if 'favicon' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['favicon']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file size (5MB limit)
        MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({'error': 'File size exceeds 5MB limit'}), 400
        
        # Check file extension
        filename = secure_filename(file.filename)
        file_ext = filename.lower().split('.')[-1] if '.' in filename else ''
        allowed_extensions = {'ico', 'png', 'jpg', 'jpeg'}
        
        if file_ext in allowed_extensions or file.content_type in ['image/x-icon', 'image/png', 'image/jpeg']:
            # Create unique filename
            import hashlib
            file_content = file.read()
            file_hash = hashlib.md5(file_content).hexdigest()[:12]
            
            # Determine file extension based on content type or filename
            if file.content_type == 'image/x-icon' or file_ext == 'ico':
                ext = 'ico'
            elif file.content_type == 'image/png' or file_ext == 'png':
                ext = 'png'
            else:
                ext = 'png'  # Default to PNG for jpg/jpeg
            
            favicon_filename = f'favicon_custom_{file_hash}.{ext}'
            favicon_path = os.path.join(DATA_DIR, favicon_filename)
            
            # Validate and process image
            try:
                from PIL import Image
                import io
                
                # Open and validate image
                image = Image.open(io.BytesIO(file_content))
                
                # Convert to appropriate format and resize if needed
                if ext == 'ico':
                    # For ICO files, keep original if it's already ICO, otherwise convert
                    if image.format != 'ICO':
                        # Resize to 32x32 for favicon
                        image = image.resize((32, 32), Image.Resampling.LANCZOS)
                        # Convert to RGBA if needed
                        if image.mode != 'RGBA':
                            image = image.convert('RGBA')
                        image.save(favicon_path, 'ICO', sizes=[(32, 32)])
                    else:
                        with open(favicon_path, 'wb') as f:
                            f.write(file_content)
                else:
                    # For PNG files
                    if image.size != (32, 32):
                        image = image.resize((32, 32), Image.Resampling.LANCZOS)
                    if image.mode != 'RGBA':
                        image = image.convert('RGBA')
                    image.save(favicon_path, 'PNG', optimize=True)
                
            except Exception as img_error:
                logger.error(f"Image processing error: {img_error}")
                return jsonify({'error': 'Invalid image file'}), 400
            
            if os.path.exists(favicon_path):
                settings = load_settings()
                settings['faviconPath'] = f'/static/favicon_custom_{file_hash}.{ext}'
                save_settings(settings)
                
                return jsonify({'success': True, 'path': settings['faviconPath']})
            else:
                return jsonify({'error': 'Failed to save file'}), 500
        else:
            return jsonify({'error': 'Invalid file type. Only ICO, PNG, JPG, JPEG allowed'}), 400
    except Exception as e:
        logger.error(f"Error uploading favicon: {e}")
        return jsonify({'error': 'Upload failed'}), 500

@app.route('/api/reset-favicon', methods=['POST'])
@require_auth
def reset_favicon():
    try:
        # Remove any custom favicon files from data directory
        import glob
        custom_favicons = glob.glob(os.path.join(DATA_DIR, 'favicon_custom*'))
        for favicon_path in custom_favicons:
            if os.path.exists(favicon_path):
                os.remove(favicon_path)
        
        settings = load_settings()
        settings['faviconPath'] = '/favicon.ico'
        save_settings(settings)
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error resetting favicon: {e}")
        return jsonify({'error': 'Reset failed'}), 500

@app.route('/api/reset-all-settings', methods=['POST'])
@require_auth
def reset_all_settings():
    try:
        # Remove any custom logo files
        import glob
        custom_logos = glob.glob(os.path.join(DATA_DIR, 'icon_custom*.png'))
        for logo_path in custom_logos:
            if os.path.exists(logo_path):
                os.remove(logo_path)
        
        save_settings(DEFAULT_SETTINGS)
        
        threading.Thread(target=restart_scheduler).start()
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error resetting all settings: {e}")
        return jsonify({'error': str(e)}), 500

def format_hire_date(date_string):
    """Format hire date to YYYY-MM-DD format"""
    if not date_string:
        return ''
    try:
        # Try parsing as ISO format first
        from datetime import datetime
        if 'T' in date_string:
            # ISO format with time
            dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d')
        else:
            # Already in date format, validate and return
            dt = datetime.strptime(date_string, '%Y-%m-%d')
            return date_string
    except (ValueError, AttributeError):
        return date_string  # Return original if parsing fails

@app.route('/api/export-xlsx')
def export_xlsx():
    """Export organizational data to XLSX format"""
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500
    
    try:
        # Load employee data
        if not os.path.exists(DATA_FILE):
            update_employee_data()
        
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        if not data:
            return jsonify({'error': 'No employee data available'}), 404
        
        # Load settings to check filtering preferences and column visibility
        settings = load_settings()
        hide_disabled_users = settings.get('hideDisabledUsers', True)
        hide_guest_users = settings.get('hideGuestUsers', True)
        hide_no_title = settings.get('hideNoTitle', True)
        ignored_departments = parse_ignored_departments(settings)
        export_column_settings = settings.get('exportXlsxColumns', {}) or {}
        is_admin = bool(session.get('authenticated'))

        column_definitions = [
            ('name', 'Name', lambda node, manager: node.get('name', '')),
            ('title', 'Title', lambda node, manager: node.get('title', '')),
            ('department', 'Department', lambda node, manager: node.get('department', '')),
            ('email', 'Email', lambda node, manager: node.get('email', '')),
            ('phone', 'Phone', lambda node, manager: node.get('phone', '')),
            ('businessPhone', 'Business Phone', lambda node, manager: node.get('businessPhone', '')),
            ('hireDate', 'Hire Date', lambda node, manager: format_hire_date(node.get('hireDate', ''))),
            ('country', 'Country', lambda node, manager: node.get('country', '')),
            ('state', 'State', lambda node, manager: node.get('state', '')),
            ('city', 'City', lambda node, manager: node.get('city', '')),
            ('office', 'Office', lambda node, manager: node.get('officeLocation', '')),
            ('manager', 'Manager', lambda node, manager: manager)
        ]

        def column_is_visible(key):
            raw_mode = export_column_settings.get(key, 'show')
            mode = str(raw_mode).lower()
            normalized_admin = mode.replace('_', '').replace('-', '')
            if mode == 'hide':
                return False
            if mode == 'admin' and not is_admin:
                return False
            if normalized_admin in {'showadminonly', 'adminonly'} and not is_admin:
                return False
            return True

        visible_columns = [col for col in column_definitions if column_is_visible(col[0])]
        if not visible_columns:
            # Always include at least the Name column to avoid empty exports
            visible_columns = [column_definitions[0]]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Organization Chart"

        # Add headers with styling
        for col_index, (_, header, _) in enumerate(visible_columns, 1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Function to flatten organizational structure
        def flatten_org_data(node, manager_name="", row_num=2):
            if not node:
                return row_num
            
            # Check if we should skip this employee based on filtering settings
            title = node.get('title', '')
            department = node.get('department', '')
            account_enabled = node.get('accountEnabled', True)
            user_type = (node.get('userType') or '').lower()
            should_skip = (hide_no_title and (not title or title.strip() == '' or title.strip() == 'No Title')) or \
                         (department_is_ignored(department, ignored_departments)) or \
                         (hide_disabled_users and not account_enabled) or \
                         (hide_guest_users and user_type == 'guest')
            
            # Add current employee only if not filtering them out
            if not should_skip:
                for col_index, (_, _, extractor) in enumerate(visible_columns, 1):
                    ws.cell(row=row_num, column=col_index, value=extractor(node, manager_name))
                row_num += 1
            
            # Add children (using current employee name as manager if not skipped, otherwise pass through current manager)
            current_manager = node.get('name', '') if not should_skip else manager_name
            for child in node.get('children', []):
                row_num = flatten_org_data(child, current_manager, row_num)
                
            return row_num
        
        # Flatten the data starting from root
        flatten_org_data(data)
        
        # Auto-adjust column widths
        for col in range(1, len(visible_columns) + 1):
            column = get_column_letter(col)
            ws.column_dimensions[column].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"org-chart-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error exporting to XLSX: {e}")
        return jsonify({'error': 'Failed to export XLSX'}), 500


def _load_missing_manager_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(MISSING_MANAGER_FILE):
            logger.info("Refreshing missing manager report cache")
            update_employee_data()

        if not os.path.exists(MISSING_MANAGER_FILE):
            logger.warning("Missing manager report cache not found after refresh")
            return []

        with open(MISSING_MANAGER_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected missing manager report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse missing manager report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading missing manager report cache: {error}")
        return []


def _load_disabled_license_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(DISABLED_LICENSE_FILE):
            logger.info("Refreshing disabled licensed users report cache")
            update_employee_data()

        if not os.path.exists(DISABLED_LICENSE_FILE):
            logger.warning("Disabled licensed users report cache not found after refresh")
            return []

        with open(DISABLED_LICENSE_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected disabled licensed users report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse disabled licensed users report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading disabled licensed users report cache: {error}")
        return []


def _load_disabled_users_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(DISABLED_USERS_FILE):
            logger.info("Refreshing disabled users report cache")
            update_employee_data()

        if not os.path.exists(DISABLED_USERS_FILE):
            logger.warning("Disabled users report cache not found after refresh")
            return []

        with open(DISABLED_USERS_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected disabled users report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse disabled users report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading disabled users report cache: {error}")
        return []


def _load_recently_disabled_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(RECENTLY_DISABLED_FILE):
            logger.info("Refreshing recently disabled employees report cache")
            update_employee_data()

        if not os.path.exists(RECENTLY_DISABLED_FILE):
            logger.warning("Recently disabled employees report cache not found after refresh")
            return []

        with open(RECENTLY_DISABLED_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected recently disabled employees report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse recently disabled employees report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading recently disabled employees report cache: {error}")
        return []


def _load_recently_hired_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(RECENTLY_HIRED_FILE):
            logger.info("Refreshing recently hired employees report cache")
            update_employee_data()

        if not os.path.exists(RECENTLY_HIRED_FILE):
            logger.warning("Recently hired employees report cache not found after refresh")
            return []

        with open(RECENTLY_HIRED_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected recently hired employees report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse recently hired employees report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading recently hired employees report cache: {error}")
        return []


def _load_last_login_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(LAST_LOGIN_FILE):
            logger.info("Refreshing last sign-in report cache")
            update_employee_data()

        if not os.path.exists(LAST_LOGIN_FILE):
            logger.warning("Last sign-in report cache not found after refresh")
            return []

        with open(LAST_LOGIN_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected last sign-in report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse last sign-in report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading last sign-in report cache: {error}")
        return []


def _load_filtered_license_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(FILTERED_LICENSE_FILE):
            logger.info("Refreshing filtered licensed users report cache")
            update_employee_data()

        if not os.path.exists(FILTERED_LICENSE_FILE):
            logger.warning("Filtered licensed users report cache not found after refresh")
            return []

        with open(FILTERED_LICENSE_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected filtered licensed users report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse filtered licensed users report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading filtered licensed users report cache: {error}")
        return []


def _load_filtered_user_data(force_refresh=False):
    try:
        if force_refresh or not os.path.exists(FILTERED_USERS_FILE):
            logger.info("Refreshing filtered users report cache")
            update_employee_data()

        if not os.path.exists(FILTERED_USERS_FILE):
            logger.warning("Filtered users report cache not found after refresh")
            return []

        with open(FILTERED_USERS_FILE, 'r') as report_file:
            data = json.load(report_file)
            if isinstance(data, list):
                return data
            logger.warning("Unexpected filtered users report format; ignoring contents")
            return []
    except json.JSONDecodeError as decode_error:
        logger.error(f"Failed to parse filtered users report cache: {decode_error}")
        return []
    except Exception as error:
        logger.error(f"Unexpected error loading filtered users report cache: {error}")
        return []


def _apply_disabled_filters(records, *, licensed_only=False, recent_days=None, include_guests=False, include_members=True):
    if recent_days is not None:
        try:
            recent_days = int(recent_days)
        except (TypeError, ValueError):
            recent_days = None

    cutoff = None
    if recent_days and recent_days > 0:
        cutoff = datetime.now(timezone.utc) - timedelta(days=recent_days)

    filtered = []

    for record in records or []:
        user_type = (record.get('userType') or '').lower()
        
        # Filter based on Azure AD userType first
        if user_type == 'guest' and not include_guests:
            continue
        if user_type == 'member' and not include_members:
            continue

        # Then filter by license status
        if licensed_only and (record.get('licenseCount') or 0) == 0:
            continue

        if cutoff is not None:
            observed = parse_graph_datetime(
                record.get('firstSeenDisabledAt')
                or record.get('disabledDate')
            )
            if not observed or observed < cutoff:
                continue

        filtered.append(record)

    return filtered


def _calculate_license_totals(records):
    return sum(record.get('licenseCount') or 0 for record in records or [])


def _apply_last_login_filters(
    records,
    *,
    include_enabled=True,
    include_disabled=True,
    include_licensed=True,
    include_unlicensed=True,
    include_members=True,
    include_guests=True,
    include_never_signed_in=True,
    inactive_days=None
):
    if not records:
        return []

    inactive_threshold = None
    require_never_signed_in = False

    if inactive_days not in (None, "", "none"):
        if isinstance(inactive_days, str) and inactive_days.lower() == 'never':
            require_never_signed_in = True
        else:
            try:
                inactive_threshold = int(inactive_days)
            except (TypeError, ValueError):
                inactive_threshold = None

    filtered = []

    for record in records:
        account_enabled = record.get('accountEnabled', True)
        if account_enabled and not include_enabled:
            continue
        if not account_enabled and not include_disabled:
            continue

        license_count = record.get('licenseCount') or 0
        if license_count > 0 and not include_licensed:
            continue
        if license_count == 0 and not include_unlicensed:
            continue

        user_type = (record.get('userType') or '').lower()
        if user_type == 'member' and not include_members:
            continue
        if user_type == 'guest' and not include_guests:
            continue

        never_signed_in = bool(record.get('neverSignedIn'))
        if never_signed_in and not include_never_signed_in:
            continue
        if require_never_signed_in and not never_signed_in:
            continue

        if inactive_threshold is not None:
            days_since = record.get('daysSinceLastActivity')
            if days_since is None or days_since < inactive_threshold:
                continue

        filtered.append(record)

    return filtered


def _apply_filtered_user_filters(
    records,
    *,
    include_enabled=True,
    include_disabled=True,
    include_licensed=True,
    include_unlicensed=True,
    include_members=True,
    include_guests=True
):
    if not records:
        return []

    filtered = []

    for record in records:
        account_enabled = record.get('accountEnabled', True)
        if account_enabled and not include_enabled:
            continue
        if not account_enabled and not include_disabled:
            continue

        license_count = record.get('licenseCount') or 0
        if license_count > 0 and not include_licensed:
            continue
        if license_count == 0 and not include_unlicensed:
            continue

        user_type = (record.get('userType') or '').lower()

        if user_type == 'guest' and not include_guests:
            continue
        if user_type == 'member' and not include_members:
            continue

        filtered.append(record)

    return filtered


def _get_disabled_records_from_request(*, force_refresh=False, apply_filters=True):
    licensed_only = request.args.get('licensedOnly', 'true').lower() == 'true'
    include_guests = request.args.get('includeGuests', 'false').lower() == 'true'
    include_members = request.args.get('includeMembers', 'true').lower() == 'true'
    recent_days_raw = request.args.get('recentDays')
    recent_days = None
    if recent_days_raw not in (None, ''):
        try:
            recent_days = int(recent_days_raw)
        except ValueError:
            logger.warning(f"Invalid recentDays value provided: {recent_days_raw}")

    records = _load_disabled_users_data(force_refresh=force_refresh)
    filtered_records = (
        _apply_disabled_filters(
            records,
            licensed_only=licensed_only,
            recent_days=recent_days,
            include_guests=include_guests,
            include_members=include_members
        )
        if apply_filters else records
    )

    filter_payload = {
        'licensedOnly': licensed_only,
        'recentDays': recent_days,
        'includeGuests': include_guests,
        'includeMembers': include_members
    }

    return filtered_records, filter_payload


@app.route('/api/reports/missing-manager')
@require_auth
def get_missing_manager_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = _load_missing_manager_data(force_refresh=refresh)
        generated_at = None
        if os.path.exists(MISSING_MANAGER_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(MISSING_MANAGER_FILE)).isoformat()

        return jsonify({
            'records': records,
            'count': len(records),
            'generatedAt': generated_at
        })
    except Exception as e:
        logger.error(f"Error loading missing manager report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/missing-manager/export')
@require_auth
def export_missing_manager_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = _load_missing_manager_data(force_refresh=refresh)

        wb = Workbook()
        ws = wb.active
        ws.title = "Missing Managers"

        headers = [
            ('name', 'Name'),
            ('title', 'Title'),
            ('department', 'Department'),
            ('email', 'Email'),
            ('businessPhone', 'Business Phone'),
            ('location', 'Location'),
            ('managerName', 'Manager Name'),
            ('reason', 'Reason')
        ]

        reason_labels = {
            'no_manager': 'No manager assigned',
            'manager_not_found': 'Manager not found in data',
            'detached': 'Detached from hierarchy'
        }

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'reason':
                    value = reason_labels.get(value, value or '')
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"missing-managers-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting missing manager report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/disabled-users')
@require_auth
def get_disabled_users_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        filtered_records, applied_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=True
        )

        generated_at = None
        if os.path.exists(DISABLED_USERS_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(DISABLED_USERS_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': applied_filters
        })
    except Exception as e:
        logger.error(f"Error loading disabled users report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/disabled-users/export')
@require_auth
def export_disabled_users_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records, _ = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=True
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Disabled Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('disabledDate', 'First Observed Disabled'),
            ('disabledDays', 'Days Since Observed Disabled'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                elif key == 'disabledDate' and value:
                    dt = parse_graph_datetime(value)
                    value = dt.date().isoformat() if dt else value
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"disabled-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting disabled users report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/disabled-this-year')
@require_auth
def get_recently_disabled_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, base_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )

        recent_days_raw = request.args.get('recentDays')
        if recent_days_raw in (None, ''):
            recent_days = 365
        else:
            try:
                recent_days = int(recent_days_raw)
            except ValueError:
                logger.warning(f"Invalid recentDays value provided for recently disabled report: {recent_days_raw}")
                recent_days = 365

        licensed_only = request.args.get('licensedOnly', 'false').lower() == 'true'
        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'

        records = _apply_disabled_filters(
            all_records,
            licensed_only=licensed_only,
            recent_days=recent_days,
            include_guests=include_guests
        )
        generated_at = None
        if os.path.exists(RECENTLY_DISABLED_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(RECENTLY_DISABLED_FILE)).isoformat()

        return jsonify({
            'records': records,
            'count': len(records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'licensedOnly': licensed_only,
                'recentDays': recent_days,
                'includeGuests': include_guests
            }
        })
    except Exception as e:
        logger.error(f"Error loading recently disabled report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/disabled-this-year/export')
@require_auth
def export_recently_disabled_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, _ = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )

        recent_days_raw = request.args.get('recentDays')
        if recent_days_raw in (None, ''):
            recent_days = 365
        else:
            try:
                recent_days = int(recent_days_raw)
            except ValueError:
                logger.warning(f"Invalid recentDays value provided for recently disabled export: {recent_days_raw}")
                recent_days = 365

        licensed_only = request.args.get('licensedOnly', 'false').lower() == 'true'
        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'

        records = _apply_disabled_filters(
            all_records,
            licensed_only=licensed_only,
            recent_days=recent_days,
            include_guests=include_guests
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Disabled Last 365 Days"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('disabledDate', 'First Observed Disabled'),
            ('disabledDays', 'Days Since Observed Disabled'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                elif key == 'disabledDate' and value:
                    dt = parse_graph_datetime(value)
                    value = dt.date().isoformat() if dt else value
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"disabled-last-365-days-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting recently disabled report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/hired-this-year')
@require_auth
def get_recently_hired_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = _load_recently_hired_data(force_refresh=refresh)
        generated_at = None
        if os.path.exists(RECENTLY_HIRED_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(RECENTLY_HIRED_FILE)).isoformat()

        return jsonify({
            'records': records,
            'count': len(records),
            'generatedAt': generated_at
        })
    except Exception as e:
        logger.error(f"Error loading recently hired report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/hired-this-year/export')
@require_auth
def export_recently_hired_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = _load_recently_hired_data(force_refresh=refresh)

        wb = Workbook()
        ws = wb.active
        ws.title = "Hired Last 365 Days"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('hireDate', 'Hire Date'),
            ('daysSinceHire', 'Days Since Hire'),
            ('managerName', 'Manager'),
            ('phone', 'Phone'),
            ('businessPhone', 'Business Phone'),
            ('location', 'Location')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'hireDate' and value:
                    value = format_hire_date(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"hired-last-365-days-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting recently hired report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


def _parse_bool_arg(value, default=True):
    if value is None:
        return default
    lowered = value.strip().lower()
    if lowered in {'true', '1', 'yes', 'on'}:
        return True
    if lowered in {'false', '0', 'no', 'off'}:
        return False
    return default


@app.route('/api/reports/last-logins')
@require_auth
def get_last_logins_report():
    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)
        include_never_signed_in = _parse_bool_arg(request.args.get('includeNeverSignedIn'), default=True)

        inactive_days_raw = request.args.get('inactiveDays')
        inactive_days = None
        if inactive_days_raw not in (None, '', 'null', 'None'):
            inactive_days = inactive_days_raw

        records = _load_last_login_data(force_refresh=refresh)
        filtered_records = _apply_last_login_filters(
            records,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests,
            include_never_signed_in=include_never_signed_in,
            inactive_days=inactive_days
        )

        generated_at = None
        if os.path.exists(LAST_LOGIN_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(LAST_LOGIN_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'licensedOnly': include_licensed and not include_unlicensed,
                'includeEnabled': include_enabled,
                'includeDisabled': include_disabled,
                'includeLicensed': include_licensed,
                'includeUnlicensed': include_unlicensed,
                'includeMembers': include_members,
                'includeGuests': include_guests,
                'includeNeverSignedIn': include_never_signed_in,
                'inactiveDays': inactive_days
            }
        })
    except Exception as error:
        logger.error(f"Error loading last sign-in report: {error}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/last-logins/export')
@require_auth
def export_last_logins_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)
        include_never_signed_in = _parse_bool_arg(request.args.get('includeNeverSignedIn'), default=True)
        inactive_days_raw = request.args.get('inactiveDays')
        inactive_days = None
        if inactive_days_raw not in (None, '', 'null', 'None'):
            inactive_days = inactive_days_raw

        records = _load_last_login_data(force_refresh=refresh)
        filtered_records = _apply_last_login_filters(
            records,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests,
            include_never_signed_in=include_never_signed_in,
            inactive_days=inactive_days
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Users by Last Sign-In"

        headers = [
            ('name', 'Name'),
            ('title', 'Title'),
            ('department', 'Department'),
            ('email', 'Email'),
            ('accountEnabled', 'Account Enabled'),
            ('userType', 'User Type'),
            ('lastActivityDate', 'Most recent sign-in'),
            ('daysSinceLastActivity', 'Days since most recent sign-in'),
            ('lastInteractiveSignIn', 'Last interactive sign-in'),
            ('daysSinceInteractiveSignIn', 'Days since interactive sign-in'),
            ('lastNonInteractiveSignIn', 'Last non-interactive sign-in'),
            ('daysSinceNonInteractiveSignIn', 'Days since non-interactive sign-in'),
            ('neverSignedIn', 'Never signed in'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(filtered_records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'accountEnabled':
                    value = 'Yes' if record.get('accountEnabled', True) else 'No'
                elif key == 'userType':
                    user_type = (record.get('userType') or '').strip()
                    value = user_type.capitalize() if user_type else ''
                elif key == 'neverSignedIn':
                    value = 'Yes' if record.get('neverSignedIn') else 'No'
                elif key == 'licenseSkus' and isinstance(value, list):
                    value = ', '.join(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 26

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"last-logins-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as error:
        logger.error(f"Error exporting last sign-in report: {error}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/disabled-licensed')
@require_auth
def get_disabled_licensed_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, base_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )

        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'
        recent_days = base_filters.get('recentDays')
        filtered_records = _apply_disabled_filters(
            all_records,
            licensed_only=True,
            recent_days=recent_days,
            include_guests=include_guests
        )
        generated_at = None
        if os.path.exists(DISABLED_LICENSE_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(DISABLED_LICENSE_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'licensedOnly': True,
                'recentDays': recent_days,
                'includeGuests': include_guests
            }
        })
    except Exception as e:
        logger.error(f"Error loading disabled licensed report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/disabled-licensed/export')
@require_auth
def export_disabled_licensed_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        all_records, base_filters = _get_disabled_records_from_request(
            force_refresh=refresh,
            apply_filters=False
        )
        recent_days = base_filters.get('recentDays')
        include_guests = request.args.get('includeGuests', 'false').lower() == 'true'
        records = _apply_disabled_filters(
            all_records,
            licensed_only=True,
            recent_days=recent_days,
            include_guests=include_guests
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Disabled Licensed Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"disabled-licensed-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting disabled licensed report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/filtered-users')
@require_auth
def get_filtered_users_report():
    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)

        if 'licensedOnly' in request.args:
            legacy_licensed_only = _parse_bool_arg(request.args.get('licensedOnly'), default=True)
            if 'includeLicensed' not in request.args:
                include_licensed = True
            if 'includeUnlicensed' not in request.args:
                include_unlicensed = not legacy_licensed_only

        records = _load_filtered_user_data(force_refresh=refresh)
        filtered_records = _apply_filtered_user_filters(
            records,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests
        )

        generated_at = None
        if os.path.exists(FILTERED_USERS_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(FILTERED_USERS_FILE)).isoformat()

        return jsonify({
            'records': filtered_records,
            'count': len(filtered_records),
            'generatedAt': generated_at,
            'appliedFilters': {
                'includeEnabled': include_enabled,
                'includeDisabled': include_disabled,
                'includeLicensed': include_licensed,
                'includeUnlicensed': include_unlicensed,
                'includeMembers': include_members,
                'includeGuests': include_guests
            }
        })
    except Exception as error:
        logger.error(f"Error loading filtered users report: {error}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/filtered-users/export')
@require_auth
def export_filtered_users_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = _parse_bool_arg(request.args.get('refresh'), default=False)

        include_enabled = _parse_bool_arg(request.args.get('includeEnabled'), default=True)
        include_disabled = _parse_bool_arg(request.args.get('includeDisabled'), default=True)
        include_licensed = _parse_bool_arg(request.args.get('includeLicensed'), default=True)
        include_unlicensed = _parse_bool_arg(request.args.get('includeUnlicensed'), default=True)
        include_members = _parse_bool_arg(request.args.get('includeMembers'), default=True)
        include_guests = _parse_bool_arg(request.args.get('includeGuests'), default=True)

        if 'licensedOnly' in request.args:
            legacy_licensed_only = _parse_bool_arg(request.args.get('licensedOnly'), default=True)
            if 'includeLicensed' not in request.args:
                include_licensed = True
            if 'includeUnlicensed' not in request.args:
                include_unlicensed = not legacy_licensed_only

        records = _load_filtered_user_data(force_refresh=refresh)
        filtered_records = _apply_filtered_user_filters(
            records,
            include_enabled=include_enabled,
            include_disabled=include_disabled,
            include_licensed=include_licensed,
            include_unlicensed=include_unlicensed,
            include_members=include_members,
            include_guests=include_guests
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('filterReasons', 'Filter Reasons'),
            ('accountEnabled', 'Account Enabled'),
            ('userType', 'User Type'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses')
        ]

        reason_labels = {
            'filter_disabled': 'Hidden: disabled user',
            'filter_guest': 'Hidden: guest account',
            'filter_no_title': 'Hidden: missing title',
            'filter_ignored_title': 'Hidden: ignored title',
            'filter_ignored_department': 'Hidden: ignored department',
            'filter_ignored_employee': 'Hidden: ignored user'
        }

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(filtered_records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'filterReasons' and isinstance(value, list):
                    converted = [reason_labels.get(reason, reason) for reason in value]
                    value = ", ".join(converted)
                elif key == 'accountEnabled':
                    value = 'Yes' if record.get('accountEnabled', True) else 'No'
                elif key == 'userType':
                    user_type = (record.get('userType') or '').strip()
                    value = user_type.capitalize() if user_type else ''
                elif key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 26

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"filtered-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as error:
        logger.error(f"Error exporting filtered users report: {error}")
        return jsonify({'error': 'Failed to export report'}), 500


@app.route('/api/reports/filtered-licensed')
@require_auth
def get_filtered_licensed_report():
    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = _load_filtered_license_data(force_refresh=refresh)
        generated_at = None
        if os.path.exists(FILTERED_LICENSE_FILE):
            generated_at = datetime.fromtimestamp(os.path.getmtime(FILTERED_LICENSE_FILE)).isoformat()

        return jsonify({
            'records': records,
            'count': len(records),
            'generatedAt': generated_at
        })
    except Exception as e:
        logger.error(f"Error loading filtered licensed report: {e}")
        return jsonify({'error': 'Failed to load report data'}), 500


@app.route('/api/reports/filtered-licensed/export')
@require_auth
def export_filtered_licensed_report():
    if not Workbook:
        return jsonify({'error': 'XLSX export not available - openpyxl not installed'}), 500

    try:
        refresh = request.args.get('refresh', 'false').lower() == 'true'
        records = _load_filtered_license_data(force_refresh=refresh)

        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Licensed Users"

        headers = [
            ('name', 'Name'),
            ('email', 'Email'),
            ('department', 'Department'),
            ('title', 'Title'),
            ('licenseCount', 'License Count'),
            ('licenseSkus', 'Licenses'),
            ('filterReasons', 'Filter Reasons')
        ]

        reason_labels = {
            'filter_disabled': 'Hidden: disabled user',
            'filter_guest': 'Hidden: guest account',
            'filter_no_title': 'Hidden: missing title',
            'filter_ignored_title': 'Hidden: ignored title',
            'filter_ignored_department': 'Hidden: ignored department',
            'filter_ignored_employee': 'Hidden: ignored user'
        }

        for column_index, (_, header_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_index, record in enumerate(records, start=2):
            for column_index, (key, _) in enumerate(headers, 1):
                value = record.get(key)
                if key == 'licenseSkus' and isinstance(value, list):
                    value = ", ".join(value)
                elif key == 'filterReasons' and isinstance(value, list):
                    converted = [reason_labels.get(reason, reason) for reason in value]
                    value = ", ".join(converted)
                ws.cell(row=row_index, column=column_index, value=value)

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"filtered-licensed-users-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting filtered licensed report: {e}")
        return jsonify({'error': 'Failed to export report'}), 500

@app.route('/api/auth-check')
def auth_check():
    """Simple endpoint to check if user is authenticated"""
    if session.get('authenticated'):
        return jsonify({'authenticated': True})
    else:
        return jsonify({'authenticated': False}), 401

@app.route('/api/search')
def search_employees():
    query = request.args.get('q', '').lower()
    
    if len(query) < 2:
        return jsonify([])
    
    try:
        if not os.path.exists(DATA_FILE):
            logger.warning(f"Data file {DATA_FILE} not found, attempting to fetch data")
            update_employee_data()
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                data = json.load(f)
        else:
            logger.error("Could not create or find employee data file")
            return jsonify([])
        
        def flatten(node, results=None):
            if results is None:
                results = []
            if node and isinstance(node, dict):
                results.append(node)
                children = node.get('children', [])
                if children and isinstance(children, list):
                    for child in children:
                        flatten(child, results)
            return results
        
        all_employees = flatten(data)
        
        results = []
        for emp in all_employees:
            if emp and isinstance(emp, dict):
                name = emp.get('name') or ''
                title = emp.get('title') or ''
                department = emp.get('department') or ''
                
                name_match = query in name.lower()
                title_match = query in title.lower()
                dept_match = query in department.lower()
                
                if name_match or title_match or dept_match:
                    results.append(emp)
        
        return jsonify(results[:10])
    except FileNotFoundError as e:
        logger.error(f"File not found in search: {e}")
        return jsonify([])
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error in search: {e}")
        return jsonify([])
    except AttributeError as e:
        logger.error(f"Attribute error in search (likely None value): {e}")
        logger.error(f"Query was: {query}")
        try:
            for emp in all_employees:
                if emp:
                    logger.debug(f"Employee data: name={emp.get('name')}, title={emp.get('title')}, dept={emp.get('department')}")
        except:
            pass
        return jsonify([])
    except Exception as e:
        logger.error(f"Error in search_employees: {e}")
        logger.error(f"Query was: {query}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/employee/<employee_id>')
def get_employee(employee_id):
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        def find_employee(node, target_id):
            if node.get('id') == target_id:
                return node
            for child in node.get('children', []):
                result = find_employee(child, target_id)
                if result:
                    return result
            return None
        
        employee = find_employee(data, employee_id)

        if employee:
            return jsonify(employee)
        return jsonify({'error': 'Employee not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/update-now', methods=['POST'])
@require_auth
@limiter.limit("1 per minute")
def trigger_update():
    try:
        threading.Thread(target=update_employee_data).start()
        logger.info(f"Manual update triggered by user: {session.get('username')}")
        return jsonify({'message': 'Update started'}), 200
    except Exception as e:
        logger.error(f"Error triggering update: {e}")
        return jsonify({'error': 'Update failed'}), 500

@app.route('/search-test')
def search_test():
    return render_template_string(get_template('search_test.html'))

@app.route('/api/debug-search')
def debug_search():
    """Debug endpoint to check search functionality"""
    try:
        info = {
            'data_file_exists': os.path.exists(DATA_FILE),
            'data_file_path': os.path.abspath(DATA_FILE) if os.path.exists(DATA_FILE) else 'Not found',
            'data_file_size': os.path.getsize(DATA_FILE) if os.path.exists(DATA_FILE) else 0,
        }
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                data = json.load(f)
                
                def count_employees(node):
                    count = 1
                    for child in node.get('children', []):
                        count += count_employees(child)
                    return count
                
                info['total_employees'] = count_employees(data) if data else 0
                info['root_employee'] = data.get('name', 'Unknown') if data else 'No data'
                info['has_children'] = bool(data.get('children')) if data else False
                
                def flatten(node, results=None):
                    if results is None:
                        results = []
                    if node and isinstance(node, dict):
                        results.append({
                            'id': node.get('id'),
                            'name': node.get('name'),
                            'title': node.get('title'),
                            'department': node.get('department')
                        })
                        children = node.get('children', [])
                        if children and isinstance(children, list):
                            for child in children:
                                flatten(child, results)
                    return results
                
                all_employees = flatten(data)
                info['sample_employees'] = all_employees[:5] if all_employees else []
                info['searchable_count'] = len(all_employees)
        else:
            info['error'] = 'Data file does not exist. Try triggering an update.'
            
        return jsonify(info)
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/force-update', methods=['POST'])
@require_auth
@limiter.limit("1 per minute")
def force_update():
    """Force an immediate update and wait for completion"""
    try:
        logger.info("Force update requested")
        update_employee_data()
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                data = json.load(f)
                
            def count_employees(node):
                if not node:
                    return 0
                count = 1
                for child in node.get('children', []):
                    count += count_employees(child)
                return count
                
            total = count_employees(data)
            return jsonify({
                'success': True,
                'message': f'Data updated successfully. {total} employees in hierarchy.',
                'file_created': True
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Update completed but no data file created. Check Azure AD credentials.',
                'file_created': False
            })
    except Exception as e:
        logger.error(f"Force update error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)